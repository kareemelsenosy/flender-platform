"""Search routes — start search, SSE progress, local + web search."""
from __future__ import annotations

import asyncio
import io
import json
import logging
import os
import threading
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List

from fastapi import APIRouter, Depends, File, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session as DBSession

from app.auth import get_current_user_id
from app.config import GOOGLE_SEARCH_KEY, GOOGLE_CSE_ID, UPLOAD_DIR
from app.core.searcher import SEARCH_CACHE_VERSION, ImageSearcher, split_and_normalize_domains
from app.services.file_safety import normalize_uploaded_name, unique_path
from app.services.ai_service import (
    ai_available,
    ai_build_search_queries,
    ai_describe_context_image,
    ai_describe_context_text,
    compose_search_instructions,
    ai_optimize_search_query,
    ai_rank_urls,
)
from app.database import SessionLocal, get_db
from app.templates_config import templates
from app.models import BrandSearchConfig, SearchCache, Session, UniqueItem

_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tiff"}
_MAX_IMAGE_UPLOAD = 500 * 1024 * 1024  # 500 MB total

router = APIRouter()
logger = logging.getLogger(__name__)

# Track active searches: session_id -> {"done": int, "total": int, "running": bool}
_search_progress: dict[int, dict] = {}

# I/O-bound search — use safe auto-scaling for large batches without exhausting upstream pools
_DEFAULT_WORKERS = 10
_MAX_WORKERS = 18


def _parse_sample_limit(value) -> int:
    try:
        limit = int(value or 0)
    except Exception:
        return 0
    return max(limit, 0)


def _resolve_search_workers(
    config: dict,
    *,
    total_groups: int,
    search_mode: str,
    use_ai: bool,
) -> int:
    try:
        requested = int(config.get("search_workers", 0) or 0)
    except Exception:
        requested = 0

    if requested > 0:
        return max(1, min(requested, _MAX_WORKERS, max(1, total_groups or 1)))

    target = _DEFAULT_WORKERS
    if search_mode == "local":
        if total_groups >= 8000:
            target = 18
        elif total_groups >= 3000:
            target = 16
        elif total_groups >= 1000:
            target = 12
    else:
        if total_groups >= 8000:
            target = 16 if use_ai else 18
        elif total_groups >= 3000:
            target = 14 if use_ai else 16
        elif total_groups >= 1000:
            target = 12 if use_ai else 14

    return max(1, min(target, _MAX_WORKERS, max(1, total_groups or 1)))


def _validate_local_folder(local_folder: str, user_id: int | None) -> str:
    if not local_folder or not user_id:
        return ""
    try:
        import pathlib
        resolved = pathlib.Path(local_folder).resolve()
        allowed_base = (UPLOAD_DIR / f"user_{user_id}").resolve()
        resolved.relative_to(allowed_base)
        return str(resolved) if resolved.is_dir() else ""
    except Exception:
        return ""


def _session_search_defaults(db: DBSession, user_id: int, session_id: int) -> dict:
    brands = [
        brand for (brand,) in db.query(UniqueItem.brand).filter(
            UniqueItem.session_id == session_id,
            UniqueItem.brand.isnot(None),
            UniqueItem.brand != "",
        ).distinct().all()
        if brand
    ]
    user_brands = db.query(BrandSearchConfig).filter(
        BrandSearchConfig.user_id == user_id
    ).all()
    searcher = ImageSearcher({
        "brand_site_urls": {bc.brand_name: bc.site_urls for bc in user_brands},
    })
    notes_by_brand = {
        bc.brand_name.strip().lower(): bc.search_notes.strip()
        for bc in user_brands
        if bc.search_notes and bc.search_notes.strip()
    }

    matched_urls: list[str] = []
    note_sections: list[str] = []
    matched_labels: list[str] = []
    for brand in brands:
        matches = searcher.matching_brand_configs(brand)
        if not matches:
            continue
        matched_labels.append(brand)
        for cfg_brand, urls in matches:
            matched_urls.extend(urls)
            note = notes_by_brand.get(cfg_brand)
            if note:
                note_sections.append(f"[{brand}]\n{note}")

    return {
        "brands": list(dict.fromkeys(brands)),
        "matched_brand_labels": list(dict.fromkeys(matched_labels)),
        "brand_urls": list(dict.fromkeys(matched_urls)),
        "search_notes": "\n\n".join(note_sections).strip(),
    }


def _reset_items_for_search(db: DBSession, session_id: int) -> int:
    """Reset retryable items so a new session-level search can process them again.

    We preserve manually-approved items, but anything that was auto-selected,
    skipped, pending, or left without a real approved URL should be searched again.
    """
    from sqlalchemy import or_

    return db.query(UniqueItem).filter(
        UniqueItem.session_id == session_id,
        or_(
            UniqueItem.review_status != "approved",
            UniqueItem.auto_selected == True,
            UniqueItem.approved_url == None,
            UniqueItem.approved_url == "",
        ),
    ).update({
        "search_status": "pending",
        "review_status": "pending",
        "approved_url": None,
        "suggested_url": None,
        "auto_selected": False,
        "search_confidence": 0.0,
        "confidence_label": "low",
        "confidence_reason": None,
        "candidates_json": "[]",
        "scores_json": "{}",
        "additional_urls_json": "[]",
    }, synchronize_session=False)


def _run_search_background(session_id: int, config: dict, user_id: int = None):
    """Run image search in background thread."""
    db = SessionLocal()
    # Snapshot the search generation at start — if it changes (remap), we abort DB updates
    search_gen = config.get("search_gen", 0)
    try:
        items = db.query(UniqueItem).filter(
            UniqueItem.session_id == session_id,
            UniqueItem.search_status == "pending",
        ).order_by(UniqueItem.id.asc()).all()

        sample_limit = _parse_sample_limit(config.get("sample_limit", 0))
        if sample_limit > 0:
            items = items[:sample_limit]

        total = len(items)
        import time as _time_mod
        _search_progress[session_id] = {
            "done": 0,
            "total": total,
            "running": True,
            "current": "",
            "started_at": _time_mod.time(),
            "workers": 0,
            "groups_total": 0,
            "groups_done": 0,
        }

        search_mode = config.get("search_mode", "web")  # web, local, both
        local_folder = config.get("local_folder", "")

        # Load user's brand search configs + notes
        brand_site_urls = {}
        brand_notes = {}  # brand_name_lower -> search instructions text
        if user_id:
            user_brands = db.query(BrandSearchConfig).filter(
                BrandSearchConfig.user_id == user_id
            ).all()
            for bc in user_brands:
                brand_site_urls[bc.brand_name.lower()] = bc.site_urls
                if bc.search_notes:
                    brand_notes[bc.brand_name.lower()] = bc.search_notes

        # Session-level search instructions (applies to all brands)
        session_notes = config.get("search_notes", "")

        # Extra brand URLs entered in Step 3 form — apply to ALL items as priority domains
        extra_brand_urls = split_and_normalize_domains(config.get("extra_brand_urls", []))

        search_config = {
            **config,
            "brand_site_urls": brand_site_urls,
            "extra_site_urls": extra_brand_urls,  # priority domains for this session
            "google_api_key": GOOGLE_SEARCH_KEY,
            "google_cse_id": GOOGLE_CSE_ID,
        }
        searcher = ImageSearcher(search_config)

        # Import local search if needed — validate path to prevent traversal (C4)
        local_search_fn = None
        if search_mode in ("local", "both") and local_folder:
            from app.services.local_search import search_local_folder
            local_folder = _validate_local_folder(local_folder, user_id)
            if local_folder:
                local_search_fn = search_local_folder
            else:
                logger.warning(f"Local folder not found or not a directory: {local_folder}")
                local_folder = ""

        use_ai = ai_available()

        grouped_items: dict[tuple[str, str, str], dict] = {}
        for item in items:
            item_dict = {
                "item_code": item.item_code,
                "color_code": item.color_code,
                "color_name": item.color_name,
                "style_name": item.style_name,
                "brand": item.brand,
                "barcode": item.barcode,
                "item_group": item.item_group,
            }
            group_identity = searcher.cache_identity(item_dict) if searcher else (
                item.item_code,
                item.color_code or "",
                (item.brand or "").lower().strip(),
            )
            grouped = grouped_items.setdefault(group_identity, {
                "items": [],
                "item_dict": item_dict,
                "label": item.item_code,
            })
            grouped["items"].append(item)

        def _search_one(item_dict: dict):
            cache_db = SessionLocal()
            try:
                cache_item_code = item_dict["item_code"]
                cache_color_code = item_dict.get("color_code") or ""
                cache_brand = (item_dict.get("brand") or "").lower()
                cached = None
                if searcher:
                    cache_item_code, cache_color_code, cache_brand = searcher.cache_identity(item_dict)
                    # Check cross-session cache first for web search only.
                    cached = cache_db.query(SearchCache).filter(
                        SearchCache.item_code == cache_item_code,
                        SearchCache.color_code == cache_color_code,
                        SearchCache.brand == cache_brand,
                        SearchCache.search_version == SEARCH_CACHE_VERSION,
                    ).first()

                if cached and cached.candidates:
                    decision = searcher.assess_match_confidence(cached.candidates, cached.scores, item_dict)
                    return cached.candidates, cached.scores, True, decision

                candidates = []
                scores = {}
                brand_label = item_dict.get("brand", "")
                matched_brand_configs = searcher.matching_brand_configs(brand_label) if searcher else []
                matched_brand_notes = [
                    brand_notes.get(cfg_brand, "")
                    for cfg_brand, _urls in matched_brand_configs
                    if brand_notes.get(cfg_brand, "").strip()
                ]
                matched_brand_urls = []
                for _cfg_brand, urls in matched_brand_configs:
                    matched_brand_urls.extend(urls)
                effective_instructions = compose_search_instructions(
                    session_notes=session_notes,
                    brand_notes=matched_brand_notes,
                    priority_domains=extra_brand_urls + matched_brand_urls,
                )

                # ── STEP 1: AI builds initial search queries ─────────────────
                # Always use AI for query building when available.
                # If user has search notes/instructions, those take priority.
                # Otherwise AI still crafts smarter queries than simple concatenation.
                ai_queries = []
                if use_ai:
                    if effective_instructions:
                        ai_queries = ai_build_search_queries(
                            item_dict, brand_label, effective_instructions
                        )
                    else:
                        # No instructions → AI still generates optimized queries
                        ai_queries = ai_optimize_search_query(
                            item_dict,
                            brand_label,
                            search_instructions=effective_instructions,
                        )

                # ── STEP 2: Local folder search ───────────────────────────────
                if local_search_fn:
                    local_results = local_search_fn(local_folder, item_dict)
                    for lr in local_results:
                        file_url = f"file://{lr['path']}"
                        candidates.append(file_url)
                        scores[file_url] = lr["score"]

                # ── STEP 3: Web search with AI queries ────────────────────────
                if search_mode == "web" or (search_mode == "both" and len(candidates) < 3):
                    web_candidates, web_scores = searcher.search(
                        item_dict, ai_queries=ai_queries or None
                    )
                    for url in web_candidates:
                        if url not in candidates:
                            candidates.append(url)
                            scores[url] = web_scores.get(url, 0)

                # ── STEP 4: AI retry if results are poor ─────────────────────
                # If we got fewer than 2 results or all scores < 0.25, AI tries
                # new queries based on what failed.
                if use_ai and search_mode != "local" and len(candidates) < 2:
                    retry_queries = ai_optimize_search_query(
                        item_dict,
                        brand_label,
                        failed_queries=ai_queries or None,
                        search_instructions=effective_instructions,
                    )
                    if retry_queries:
                        retry_candidates, retry_scores = searcher.search(
                            item_dict, ai_queries=retry_queries
                        )
                        for url in retry_candidates:
                            if url not in candidates:
                                candidates.append(url)
                                scores[url] = retry_scores.get(url, 0)

                # ── STEP 5: AI re-ranks the final URL list ────────────────────
                # Only web URLs go through AI ranking (local file:// paths are kept as-is)
                if use_ai and candidates:
                    web_urls = [u for u in candidates if not u.startswith("file://")]
                    local_urls = [u for u in candidates if u.startswith("file://")]
                    if web_urls:
                        ai_primary_mode = bool(searcher and searcher.should_force_ai_primary(item_dict))
                        ranked_web = ai_rank_urls(
                            web_urls,
                            item_dict,
                            brand_label,
                            scores=scores,
                            prefer_vision=ai_primary_mode,
                        )
                        # Rebuild candidates: local first, then AI-ranked web
                        reranked = local_urls + ranked_web
                        # In strict/vision-heavy categories we let AI order drive the
                        # final score more strongly so clean packshots win decisively.
                        new_scores = {}
                        for i, url in enumerate(reranked):
                            base = scores.get(url, 0.5)
                            if ai_primary_mode and not url.startswith("file://"):
                                ai_target = max(0.45, 0.97 - i * 0.09)
                                new_scores[url] = min(round((base * 0.4) + (ai_target * 0.6), 2), 1.0)
                            else:
                                position_bonus = max(0.0, 0.1 - i * 0.02)
                                new_scores[url] = min(round(base + position_bonus, 2), 1.0)
                        candidates = reranked
                        scores = new_scores

                # ── STEP 6: Save to cache — web search only ──────────────────
                decision = searcher.assess_match_confidence(candidates, scores, item_dict)

                if search_mode != "local":
                    try:
                        existing_cache = cache_db.query(SearchCache).filter(
                            SearchCache.item_code == cache_item_code,
                            SearchCache.color_code == cache_color_code,
                            SearchCache.brand == cache_brand,
                            SearchCache.search_version == SEARCH_CACHE_VERSION,
                        ).first()
                        if existing_cache:
                            existing_cache.candidates = candidates
                            existing_cache.scores = scores
                        else:
                            new_cache = SearchCache(
                                item_code=cache_item_code,
                                color_code=cache_color_code,
                                brand=cache_brand,
                                search_version=SEARCH_CACHE_VERSION,
                            )
                            new_cache.candidates = candidates
                            new_cache.scores = scores
                            cache_db.add(new_cache)
                        cache_db.commit()
                    except Exception:
                        cache_db.rollback()

                return candidates, scores, False, decision
            finally:
                cache_db.close()

        group_entries = list(grouped_items.values())
        workers = _resolve_search_workers(
            config,
            total_groups=len(group_entries),
            search_mode=search_mode,
            use_ai=use_ai,
        )
        _search_progress[session_id]["workers"] = workers
        _search_progress[session_id]["groups_total"] = len(group_entries)

        # Collect all results first so we can dedup `approved_url` across
        # different base-item-codes once every search has completed. Different
        # item codes MUST NOT end up with the same picture.
        completed_results: list[tuple[dict, list[str], dict[str, float], dict[str, object]]] = []

        with ThreadPoolExecutor(max_workers=max(1, min(workers, len(group_entries) or 1))) as executor:
            futures = {}
            for grouped in group_entries:
                futures[executor.submit(_search_one, grouped["item_dict"])] = grouped

            for future in as_completed(futures):
                # Check if search was cancelled (remap bumped search_gen)
                try:
                    current_sess = db.get(Session, session_id)
                    current_gen = (current_sess.config or {}).get("search_gen", 0) if current_sess else -1
                    if current_gen != search_gen:
                        logger.info(f"Search cancelled for session {session_id} (gen {search_gen} != {current_gen})")
                        break
                except Exception:
                    pass

                grouped = futures[future]
                try:
                    candidates, scores, _from_cache, decision = future.result()
                    completed_results.append((grouped, candidates, scores, decision))
                except Exception as e:
                    logger.error(f"Search error: {e}")

                _search_progress[session_id]["done"] += len(grouped["items"])
                _search_progress[session_id]["groups_done"] = len(completed_results)
                label = grouped["label"]
                if len(grouped["items"]) > 1:
                    label = f"{label} (+{len(grouped['items']) - 1})"
                _search_progress[session_id]["current"] = label

        # Dedup: give the highest-confidence group first pick of its best URL,
        # then later groups must pick a URL not yet claimed by another group.
        completed_results.sort(
            key=lambda gr: float((gr[3] or {}).get("score", 0.0) or 0.0),
            reverse=True,
        )
        claimed_urls: set[str] = set()
        for grouped, candidates, scores, _decision in completed_results:
            sorted_urls = sorted(candidates, key=lambda u: scores.get(u, 0.0), reverse=True)
            chosen = ""
            for url in sorted_urls:
                if url and url not in claimed_urls:
                    chosen = url
                    break
            if chosen:
                claimed_urls.add(chosen)

            # Reorder candidates so the item's own chosen URL is first, then
            # any URLs no other group has claimed, then already-claimed URLs
            # (still useful as manual overrides in the review UI).
            free = [u for u in sorted_urls if u == chosen or u not in claimed_urls]
            taken = [u for u in sorted_urls if u != chosen and u in claimed_urls]
            reordered = [u for u in ([chosen] if chosen else []) + free + taken]
            seen: set[str] = set()
            final_candidates: list[str] = []
            for url in reordered:
                if url and url not in seen:
                    seen.add(url)
                    final_candidates.append(url)

            final_decision = searcher.assess_match_confidence(
                final_candidates,
                scores,
                grouped["item_dict"],
                prefer_first=True,
            )
            final_label = str(final_decision.get("label") or "low")
            final_score = float(final_decision.get("score", 0.0) or 0.0)
            final_reason = str(final_decision.get("reason") or "").strip() or None
            final_suggested = str(final_decision.get("suggested_url") or chosen or "").strip() or None
            auto_approve = bool(final_decision.get("auto_approve")) and bool(final_suggested)

            for source_item in grouped["items"]:
                db_item = db.get(UniqueItem, source_item.id)
                if not db_item:
                    continue
                db_item.candidates = final_candidates
                db_item.scores = scores
                db_item.search_status = "done"
                db_item.suggested_url = final_suggested
                db_item.search_confidence = final_score
                db_item.confidence_label = final_label
                db_item.confidence_reason = final_reason
                if auto_approve:
                    db_item.approved_url = final_suggested
                    db_item.review_status = "approved"
                    db_item.auto_selected = True
                else:
                    db_item.approved_url = None
                    db_item.review_status = "pending"
                    db_item.auto_selected = False
        db.commit()

        # Update session status — only if this search generation is still current
        # (guards against a remap happening mid-search that bumped search_gen)
        sess = db.get(Session, session_id)
        current_gen = (sess.config or {}).get("search_gen", 0) if sess else -1
        if sess and current_gen == search_gen:
            sess.status = "reviewing"
            sess.searched_items = db.query(UniqueItem).filter(
                UniqueItem.session_id == session_id,
                UniqueItem.search_status == "done",
            ).count()
            db.commit()

        _search_progress[session_id]["running"] = False

        # L3: Schedule cleanup of progress entry after 5 minutes (gives SSE time to drain)
        import time as _time
        def _cleanup_progress():
            _time.sleep(300)
            _search_progress.pop(session_id, None)
        threading.Thread(target=_cleanup_progress, daemon=True).start()

        # Notify user
        if user_id:
            from app.services.notifications import add_notification
            sess_name = sess.name if sess else f"Session #{session_id}"
            add_notification(
                user_id, "search_done",
                "Image Search Complete",
                f"{total} items searched — {sess_name}",
                session_id,
                [
                    {"label": "Review Images", "url": f"/review/{session_id}"},
                    {"label": "Export", "url": f"/generate/{session_id}"},
                ],
            )
    except Exception as e:
        logger.error(f"Background search failed: {e}")
        prog = _search_progress.get(session_id)
        if prog:
            prog["running"] = False
        else:
            _search_progress[session_id] = {
                "done": 0, "total": 0, "running": False, "current": "",
            }
    finally:
        db.close()


# ── Batch search: start multiple sessions in parallel ─────────────────────────
# NOTE: these static routes must come before /search/{session_id} dynamic routes.

@router.post("/search/batch/start")
async def start_batch_search(request: Request, db: DBSession = Depends(get_db)):
    """Start image search for multiple sessions simultaneously."""
    uid = get_current_user_id(request)
    if not uid:
        return JSONResponse({"error": "unauthorized"}, status_code=401)

    data = await request.json()
    session_ids: list[int] = data.get("session_ids", [])
    search_mode = data.get("search_mode", "web")
    local_folder = _validate_local_folder(data.get("local_folder", ""), uid)
    brand_urls = split_and_normalize_domains(data.get("brand_urls", []))
    search_notes = str(data.get("search_notes", "") or "").strip()
    try:
        search_workers = int(data.get("search_workers", 0) or 0)
    except Exception:
        search_workers = 0

    if not session_ids:
        return JSONResponse({"error": "No session IDs provided"}, status_code=400)
    if search_mode == "local" and not local_folder:
        return JSONResponse({"error": "Upload an image folder before starting local search"}, status_code=400)

    started = []
    for sid in session_ids:
        sess = db.query(Session).filter(Session.id == sid, Session.user_id == uid).first()
        if not sess:
            continue
        if _search_progress.get(sid, {}).get("running"):
            started.append(sid)
            continue

        config = dict(sess.config or {})
        config["search_mode"] = search_mode
        config["local_folder"] = local_folder
        config["search_notes"] = search_notes
        config["extra_brand_urls"] = brand_urls
        if search_workers > 0:
            config["search_workers"] = search_workers
        else:
            config.pop("search_workers", None)
        config["search_gen"] = config.get("search_gen", 0) + 1
        reset_count = _reset_items_for_search(db, sid)
        sess.config = config
        sess.searched_items = 0
        sess.status = "searching" if reset_count > 0 else "reviewing"
        db.commit()

        if reset_count == 0:
            continue

        thread = threading.Thread(
            target=_run_search_background,
            args=(sid, config, uid),
            daemon=True,
        )
        thread.start()
        started.append(sid)

    return JSONResponse({"ok": True, "started": started})


@router.get("/search/batch/progress")
async def batch_search_progress_sse(session_ids: str, request: Request, db: DBSession = Depends(get_db)):
    """SSE for multiple sessions' search progress. Pass session_ids as comma-separated."""
    uid = get_current_user_id(request)
    if not uid:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    ids = [int(x) for x in session_ids.split(",") if x.strip().isdigit()]
    if not ids:
        return JSONResponse({"error": "No session IDs provided"}, status_code=400)
    owned_count = db.query(Session.id).filter(
        Session.user_id == uid,
        Session.id.in_(ids),
    ).count()
    if owned_count != len(set(ids)):
        return JSONResponse({"error": "forbidden"}, status_code=403)

    async def event_stream():
        # Give threads up to 3 s to register themselves in _search_progress
        for _ in range(30):
            if all(sid in _search_progress for sid in ids):
                break
            await asyncio.sleep(0.1)

        while True:
            snapshot = {}
            all_done = True
            for sid in ids:
                p = _search_progress.get(sid)
                if p is None:
                    snapshot[str(sid)] = {"done": 0, "total": 0, "running": True, "current": ""}
                    all_done = False
                else:
                    snapshot[str(sid)] = p
                    if p.get("running") or (p.get("total", 0) > 0 and p.get("done", 0) < p.get("total", 0)):
                        all_done = False

            yield f"data: {json.dumps(snapshot)}\n\n"
            if all_done:
                yield f"data: {json.dumps({**snapshot, 'complete': True})}\n\n"
                break
            await asyncio.sleep(0.5)

    return StreamingResponse(event_stream(), media_type="text/event-stream")


# ── Image upload for local search ────────────────────────────────────────────

@router.post("/search/{session_id}/upload-images")
async def upload_images_for_search(
    session_id: int,
    request: Request,
    files: List[UploadFile] = File(...),
    db: DBSession = Depends(get_db),
):
    """
    Accept image files (JPG/PNG/WebP/etc. or ZIP) uploaded from the user's browser.
    Stores them server-side so the local search can run against them.
    Returns {"ok": True, "folder_path": "...", "image_count": N}.
    """
    uid = get_current_user_id(request)
    if not uid:
        return JSONResponse({"error": "unauthorized"}, status_code=401)

    sess = db.query(Session).filter(Session.id == session_id, Session.user_id == uid).first()
    if not sess:
        return JSONResponse({"error": "not found"}, status_code=404)

    img_dir = UPLOAD_DIR / f"user_{uid}" / f"session_{session_id}_images"
    img_dir.mkdir(parents=True, exist_ok=True)

    image_count = 0
    total_bytes = 0
    extracted_bytes = 0

    for upload in files:
        content = await upload.read()
        total_bytes += len(content)
        if total_bytes > _MAX_IMAGE_UPLOAD:
            return JSONResponse({"error": "Total upload size exceeds 500 MB limit"}, status_code=413)

        _display_name, safe_name = normalize_uploaded_name(upload.filename or "image", default="image")
        ext = os.path.splitext(safe_name)[1].lower()

        if ext == ".zip":
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as zf:
                    for info in zf.infolist():
                        if info.is_dir():
                            continue
                        _zip_display, zip_safe_name = normalize_uploaded_name(info.filename, default="image")
                        zext = os.path.splitext(zip_safe_name)[1].lower()
                        if zext not in _IMAGE_EXTENSIONS:
                            continue
                        extracted_bytes += info.file_size
                        if extracted_bytes > _MAX_IMAGE_UPLOAD:
                            return JSONResponse({"error": "Extracted image size exceeds 500 MB limit"}, status_code=413)
                        dest = unique_path(img_dir, zip_safe_name)
                        dest.write_bytes(zf.read(info))
                        image_count += 1
            except zipfile.BadZipFile:
                logger.warning(f"Skipping bad ZIP: {safe_name}")
        elif ext in _IMAGE_EXTENSIONS:
            dest = unique_path(img_dir, safe_name)
            dest.write_bytes(content)
            image_count += 1

    logger.info(f"Session {session_id}: uploaded {image_count} images to {img_dir}")
    return JSONResponse({"ok": True, "folder_path": str(img_dir), "image_count": image_count})


_CONTEXT_TEXT_EXTS = {".txt", ".md", ".csv", ".tsv", ".json", ".html", ".htm"}
_CONTEXT_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}
_CONTEXT_SPREADSHEET_EXTS = {".xlsx", ".xls"}
_MAX_CONTEXT_FILE = 20 * 1024 * 1024  # 20 MB


def _extract_spreadsheet_text(content: bytes, ext: str) -> str:
    """Read rows from an xlsx/xls file and return a compact text preview."""
    try:
        if ext == ".xlsx":
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            try:
                parts: list[str] = []
                for ws in wb.worksheets[:4]:
                    rows: list[str] = []
                    for row in ws.iter_rows(values_only=True):
                        values = [str(cell or "").replace("\n", " ").strip() for cell in row[:40]]
                        if not any(values):
                            continue
                        rows.append(",".join(value.replace(",", ";") for value in values))
                        if len(rows) >= 80:
                            break
                    if rows:
                        parts.append(f"# Sheet: {ws.title}\n" + "\n".join(rows))
                return "\n\n".join(parts)
            finally:
                wb.close()

        import pandas as pd
        xls = pd.ExcelFile(io.BytesIO(content))
        parts: list[str] = []
        for sheet_name in xls.sheet_names[:4]:
            df = xls.parse(sheet_name, nrows=80)
            parts.append(f"# Sheet: {sheet_name}\n{df.to_csv(index=False)}")
        return "\n\n".join(parts)
    except Exception as exc:
        logger.warning(f"context-file: spreadsheet extract failed: {exc}")
        return ""


def _extract_pdf_text(content: bytes) -> str:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        try:
            from PyPDF2 import PdfReader  # type: ignore
        except Exception:
            return ""
    try:
        reader = PdfReader(io.BytesIO(content))
        chunks: list[str] = []
        for page in reader.pages[:15]:
            try:
                chunks.append(page.extract_text() or "")
            except Exception:
                continue
        return "\n".join(chunks)
    except Exception as exc:
        logger.warning(f"context-file: pdf extract failed: {exc}")
        return ""


def _mime_for_ext(ext: str) -> str:
    return {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
        ".gif": "image/gif",
        ".bmp": "image/bmp",
    }.get(ext, "image/jpeg")


@router.post("/search/{session_id}/describe-context")
async def describe_context_file(
    session_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: DBSession = Depends(get_db),
):
    """Accept a user-supplied reference file (image/pdf/xlsx/txt) that explains
    what items we're searching for. Returns an AI summary the caller can paste
    into the AI Search Instructions textarea."""
    uid = get_current_user_id(request)
    if not uid:
        return JSONResponse({"error": "unauthorized"}, status_code=401)

    sess = db.query(Session).filter(Session.id == session_id, Session.user_id == uid).first()
    if not sess:
        return JSONResponse({"error": "not found"}, status_code=404)

    if not ai_available():
        return JSONResponse({"error": "AI service not configured"}, status_code=503)

    content = await file.read()
    if not content:
        return JSONResponse({"error": "Empty file"}, status_code=400)
    if len(content) > _MAX_CONTEXT_FILE:
        return JSONResponse({"error": "File exceeds 20 MB limit"}, status_code=413)

    _display_name, safe_name = normalize_uploaded_name(file.filename or "context", default="context")
    ext = os.path.splitext(safe_name)[1].lower()

    summary: str | None = None
    source_kind = "unknown"

    try:
        if ext in _CONTEXT_IMAGE_EXTS:
            source_kind = "image"
            summary = ai_describe_context_image(content, mime_type=_mime_for_ext(ext))
        elif ext == ".pdf":
            source_kind = "pdf"
            text = _extract_pdf_text(content)
            if text.strip():
                summary = ai_describe_context_text(text, filename=file.filename or "context.pdf")
            else:
                return JSONResponse({
                    "error": "Could not read text from this PDF. Try converting it to an image or text file first.",
                }, status_code=415)
        elif ext in _CONTEXT_SPREADSHEET_EXTS:
            source_kind = "spreadsheet"
            text = _extract_spreadsheet_text(content, ext)
            if text.strip():
                summary = ai_describe_context_text(text, filename=file.filename or "context.xlsx")
        elif ext in _CONTEXT_TEXT_EXTS or not ext:
            source_kind = "text"
            try:
                text = content.decode("utf-8", errors="replace")
            except Exception:
                text = ""
            if text.strip():
                summary = ai_describe_context_text(text, filename=file.filename or "context.txt")
        else:
            return JSONResponse({
                "error": f"Unsupported file type '{ext}'. Use an image, PDF, spreadsheet, or text file.",
            }, status_code=415)
    except Exception as exc:
        logger.exception(f"describe-context failed for session {session_id}: {exc}")
        return JSONResponse({"error": "Failed to analyze the file."}, status_code=500)

    summary = (summary or "").strip()
    if not summary:
        return JSONResponse({"error": "AI did not return a description. Try a different file."}, status_code=502)

    logger.info(f"Session {session_id}: generated context summary from {source_kind} ({len(summary)} chars)")
    return JSONResponse({"ok": True, "description": summary, "source": source_kind, "filename": file.filename})


# ── Per-session routes ────────────────────────────────────────────────────────

@router.get("/search/{session_id}", response_class=HTMLResponse)
def search_page(session_id: int, request: Request, db: DBSession = Depends(get_db)):
    uid = get_current_user_id(request)
    if not uid:
        return RedirectResponse("/login", status_code=302)

    sess = db.query(Session).filter(Session.id == session_id, Session.user_id == uid).first()
    if not sess:
        return RedirectResponse("/", status_code=302)

    pending_count = db.query(UniqueItem).filter(
        UniqueItem.session_id == session_id,
        UniqueItem.search_status == "pending",
    ).count()
    done_count = db.query(UniqueItem).filter(
        UniqueItem.session_id == session_id,
        UniqueItem.search_status == "done",
    ).count()

    # If fully done searching, go to review. If some items are still pending,
    # keep Step 3 accessible so users can test a subset first, then continue.
    if sess.status in ("reviewing", "completed") and pending_count == 0:
        return RedirectResponse(f"/review/{session_id}", status_code=302)

    # Check if search is already running
    prog = _search_progress.get(session_id, {})
    is_running = prog.get("running", False)

    # If session says "searching" but no active background thread,
    # the search finished (or crashed) — check if items were searched
    if sess.status == "searching" and not is_running:
        if done_count > 0:
            sess.status = "reviewing"
            db.commit()
            if pending_count == 0:
                return RedirectResponse(f"/review/{session_id}", status_code=302)

    defaults = _session_search_defaults(db, uid, session_id)
    config = dict(sess.config or {})

    return templates.TemplateResponse(request, "search.html", {
        "session": sess,
        "is_running": is_running,
        "current_search_mode": config.get("search_mode", "web"),
        "default_brand_urls": split_and_normalize_domains(config.get("extra_brand_urls", [])) or defaults["brand_urls"],
        "default_search_notes": str(config.get("search_notes", "") or "").strip() or defaults["search_notes"],
        "default_sample_limit": _parse_sample_limit(config.get("sample_limit", 50)) or 50,
        "matched_brand_labels": defaults["matched_brand_labels"],
        "pending_search_count": pending_count,
        "searched_item_count": done_count,
    })


@router.post("/search/{session_id}/start")
async def start_search(session_id: int, request: Request, db: DBSession = Depends(get_db)):
    """Start search with user-selected options."""
    uid = get_current_user_id(request)
    if not uid:
        return JSONResponse({"error": "unauthorized"}, status_code=401)

    sess = db.query(Session).filter(Session.id == session_id, Session.user_id == uid).first()
    if not sess:
        return JSONResponse({"error": "not found"}, status_code=404)

    data = await request.json()
    search_mode = data.get("search_mode", "web")  # web, local, both
    local_folder = _validate_local_folder(data.get("local_folder", ""), uid)
    brand_urls = split_and_normalize_domains(data.get("brand_urls", []))  # Additional brand URLs for this search
    search_notes = str(data.get("search_notes", "") or "").strip()
    sample_limit = _parse_sample_limit(data.get("sample_limit", 0))
    try:
        search_workers = int(data.get("search_workers", 0) or 0)
    except Exception:
        search_workers = 0
    if search_mode == "local" and not local_folder:
        return JSONResponse({"error": "Upload an image folder before starting local search"}, status_code=400)

    pending_count = db.query(UniqueItem).filter(
        UniqueItem.session_id == session_id,
        UniqueItem.search_status == "pending",
    ).count()
    done_count = db.query(UniqueItem).filter(
        UniqueItem.session_id == session_id,
        UniqueItem.search_status == "done",
    ).count()
    continuing_partial = pending_count > 0 and done_count > 0 and sess.status in ("reviewing", "completed")

    # Update session config with search settings
    config = dict(sess.config or {})
    config["search_mode"] = search_mode
    config["local_folder"] = local_folder
    config["extra_brand_urls"] = brand_urls
    config["search_notes"] = search_notes
    if sample_limit > 0:
        config["sample_limit"] = sample_limit
    else:
        config.pop("sample_limit", None)
    if search_workers > 0:
        config["search_workers"] = search_workers
    else:
        config.pop("search_workers", None)
    # Bump search generation so any stale background threads won't overwrite status
    config["search_gen"] = config.get("search_gen", 0) + 1
    reset_count = pending_count if continuing_partial else _reset_items_for_search(db, session_id)
    sess.config = config
    sess.searched_items = done_count if continuing_partial else 0
    sess.status = "searching" if reset_count > 0 else "reviewing"
    db.commit()

    # Clear any stale progress so old threads don't block a new search from starting
    _search_progress.pop(session_id, None)

    if reset_count == 0:
        return JSONResponse({
            "ok": True,
            "redirect": f"/review/{session_id}",
            "message": "Nothing left to search",
        })

    # Start search if not already running
    if session_id not in _search_progress or not _search_progress.get(session_id, {}).get("running"):
        thread = threading.Thread(
            target=_run_search_background,
            args=(session_id, config, uid),
            daemon=True,
        )
        thread.start()

    return JSONResponse({"ok": True})


@router.get("/search/{session_id}/progress")
async def search_progress_sse(session_id: int, request: Request, db: DBSession = Depends(get_db)):
    """SSE endpoint for real-time search progress."""
    uid = get_current_user_id(request)
    if not uid:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    sess = db.query(Session).filter(Session.id == session_id, Session.user_id == uid).first()
    if not sess:
        return JSONResponse({"error": "forbidden"}, status_code=403)

    async def event_stream():
        while True:
            progress = _search_progress.get(session_id, {"done": 0, "total": 0, "running": False, "current": ""})
            data = json.dumps(progress)
            yield f"data: {data}\n\n"

            if not progress.get("running") and progress.get("done", 0) >= progress.get("total", 1):
                yield f"data: {json.dumps({**progress, 'complete': True})}\n\n"
                break
            await asyncio.sleep(0.5)

    return StreamingResponse(event_stream(), media_type="text/event-stream")
