"""
Core generator — Excel order sheet with embedded images.
Supports both simple ordersheet and full 23-column standard format with
product groups, merged cells, and Row helper columns.
"""
from __future__ import annotations

import hashlib
import io
import logging
import os
import re
import shutil
import tempfile
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from app.services.file_safety import normalize_folder_name

# ─── Styles ────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
SUMMARY_FILL = PatternFill("solid", fgColor="F1F5F9")

GROUP_FILLS = [
    PatternFill("solid", fgColor="EFF6FF"),  # light blue
    PatternFill("solid", fgColor="F9FAFB"),  # light gray
]
PRICE_FILLS = [
    PatternFill("solid", fgColor="DBEAFE"),  # blue tint
    PatternFill("solid", fgColor="F3F4F6"),  # gray tint
]
QTY_FILL = PatternFill("solid", fgColor="FEF9C3")       # yellow = editable
TOTAL_FILL = PatternFill("solid", fgColor="DCFCE7")     # green = calculated
FREESTOCK_FILL = PatternFill("solid", fgColor="D1FAE5")  # light green
PICTURE_FILL = PatternFill("solid", fgColor="D0D0D0")   # neutral gray — visible bg for dark products

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

BODY_FONT = Font(size=9, name="Calibri")
PRICE_FONT = Font(bold=True, size=9, color="1D4ED8", name="Calibri")
LINK_FONT = Font(size=9, color="2563EB", underline="single", name="Calibri")
FORMULA_FONT = Font(bold=True, size=9, color="166534", name="Calibri")
EDITABLE_FONT = Font(size=9, color="92400E", name="Calibri")
SUMMARY_QTY_FONT = Font(bold=True, size=10, color="92400E", name="Calibri")
SUMMARY_TOTAL_FONT = Font(bold=True, size=10, color="166534", name="Calibri")

_DL_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
}

logger = logging.getLogger(__name__)

IMAGE_PX = 150
IMAGE_PT = IMAGE_PX * 0.75
TEXT_ROW_H = 22


def _best_image_url_for_group(items: list[dict], start: int, end: int) -> str:
    """Pick the first usable image URL across rows in a product group.

    Falls back through approved_url → suggested_url → pictures_url so an item
    still gets a picture in the export even when only one of those is set
    (e.g. legacy items that hold the URL on `pictures_url`).
    """
    for key in ("approved_url", "suggested_url", "pictures_url"):
        for i in range(start, end + 1):
            url = items[i].get(key) or ""
            if not isinstance(url, str):
                continue
            url = url.strip()
            if not url:
                continue
            # `pictures_url` is often a Dropbox folder link — only embed it if
            # it looks directly downloadable (the Dropbox-to-direct conversion
            # for /scl/fi/ links happens later in _download_image).
            if key == "pictures_url" and not (
                url.endswith((".jpg", ".jpeg", ".png", ".webp", ".gif"))
                or "dropboxusercontent.com" in url
                or "dl=1" in url
            ):
                continue
            return url
    return ""


def detect_currency_symbol(items: list[dict], currency_override: str = "") -> str:
    """Detect currency from an explicit override or from WHS Price sample values."""
    if currency_override:
        return currency_override
    # Check first few items' wholesale_price raw string representation
    for item in items[:10]:
        whs = str(item.get("wholesale_price_raw", item.get("wholesale_price", "")))
        wu = whs.upper()
        if "AED" in wu or wu.startswith("DH"):
            return "AED "
        if wu.startswith("$") or "USD" in wu:
            return "$"
        if wu.startswith("£") or "GBP" in wu:
            return "£"
        if wu.startswith("€") or "EUR" in wu:
            return "€"
    return "€"

# Standard ordersheet columns
STANDARD_COLUMNS = {
    "Picture":               {"width": 21,  "align": CENTER},
    "Brand Name":            {"width": 18,  "align": LEFT},
    "Item Group":            {"width": 16,  "align": LEFT},
    "Manufacturer Code":     {"width": 20,  "align": LEFT},
    "Web Description 2":     {"width": 32,  "align": LEFT},
    "Barcode":               {"width": 16,  "align": CENTER},
    "Gender":                {"width": 10,  "align": CENTER},
    "Color":                 {"width": 14,  "align": LEFT},
    "Size":                  {"width": 8,   "align": CENTER},
    "Stock":                 {"width": 11,  "align": CENTER},
    "Comming Soon":          {"width": 12,  "align": CENTER},
    "QTY":                   {"width": 8,   "align": CENTER},
    "QTY Total":             {"width": 14,  "align": RIGHT},
    "WHS Price":             {"width": 12,  "align": RIGHT},
    "RRP Price":             {"width": 12,  "align": RIGHT},
    "Pictures":              {"width": 10,  "align": CENTER},
    # Row helper columns (grouped/hidden)
    "Row WHS Price":         {"width": 0},
    "Row RRP Price":         {"width": 0},
    "Row Manufacturer Code": {"width": 0},
    "Row Web Description 2": {"width": 0},
    "Row Color":             {"width": 0},
    "Row Brand Name":        {"width": 0},
    "ItemCode":              {"width": 14,  "align": LEFT},
}


def _has_comming_soon_column(items: list[dict]) -> bool:
    for item in items:
        if "comming_soon_qty" not in item:
            continue
        value = item.get("comming_soon_qty")
        if value is None:
            continue
        if str(value).strip() != "":
            return True
    return False


def _coerce_sheet_value(value: Any) -> Any:
    text = str(value or "").strip()
    if not text:
        return ""
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.\d+", text):
        return float(text)
    return text


class OrderSheetGenerator:
    """Generate Excel order sheets with embedded images."""

    def __init__(self, config: dict[str, Any] | None = None,
                 progress_callback: Any = None):
        cfg = config or {}
        self.img_size = tuple(cfg.get("image_size", [150, 150]))
        self.row_height_pt = int(cfg.get("row_height_px", 100)) * 0.75
        self.save_images = cfg.get("save_images_to_folder", False)
        self._progress = progress_callback  # callable(downloaded, total, stage)

    def _report(self, downloaded: int, total: int, stage: str):
        if self._progress:
            self._progress(downloaded, total, stage)

    def generate(self, items: list[dict], output_dir: str,
                 input_filename: str = "export", brand: str = "",
                 currency: str = "", google_sheet_tabs: list[str] | None = None) -> str:
        """Generate standard ordersheet Excel from approved items."""
        os.makedirs(output_dir, exist_ok=True)

        today = date.today().strftime("%Y%m%d")
        safe_brand = re.sub(r"[^\w\-]", "_", brand) if brand else "FLENDER"
        safe_input = re.sub(r"[^\w\-]", "_", os.path.splitext(input_filename)[0])
        out_filename = f"{today}_{safe_brand}_{safe_input}_OrderSheet.xlsx"
        out_path = os.path.join(output_dir, out_filename)

        # Keep Dropbox's staging folder exactly as requested. Product folders
        # inside it use spaces, while image filenames keep underscores.
        images_dir = None
        if self.save_images:
            images_dir = os.path.join(output_dir, "images", "_READY TO UPDATE")
            # Re-exporting a session reuses the same output directory. Clear the
            # image staging folder first so old underscore-style folders/files
            # never get mixed into the fresh ZIP.
            if os.path.isdir(images_dir):
                shutil.rmtree(images_dir)
            os.makedirs(images_dir, exist_ok=True)

        wb = Workbook()
        sheet_groups = self._build_sheet_groups(items, google_sheet_tabs or [])
        progress_state = {
            "downloaded": 0,
            "total": sum(self._count_unique_sheet_urls(sheet_items, images_dir) for _title, sheet_items in sheet_groups),
        }
        self._report(0, progress_state["total"], "downloading")

        used_sheet_names: set[str] = set()
        tmp_images: list[str] = []
        active_sheet = wb.active

        for idx, (sheet_title, sheet_items) in enumerate(sheet_groups):
            ws = active_sheet if idx == 0 else wb.create_sheet()
            ws.title = self._make_sheet_name(sheet_title, used_sheet_names)
            self._populate_worksheet(
                wb=wb,
                ws=ws,
                items=sheet_items,
                images_dir=images_dir,
                currency=currency,
                progress_state=progress_state,
                tmp_images=tmp_images,
            )

        self._report(progress_state["total"], progress_state["total"], "saving")
        wb.save(out_path)
        self._report(progress_state["total"], progress_state["total"], "done")

        # Clean up temp images and the per-export download spill dir.
        for p in tmp_images:
            try:
                if os.path.isdir(p):
                    shutil.rmtree(p, ignore_errors=True)
                else:
                    os.remove(p)
            except Exception:
                pass

        return os.path.abspath(out_path)

    def _build_sheet_groups(self, items: list[dict], google_sheet_tabs: list[str]) -> list[tuple[str, list[dict]]]:
        if not google_sheet_tabs:
            return [("Order Sheet", items)]

        items_by_sheet: dict[str, list[dict]] = {}
        for item in items:
            sheet_name = str(item.get("source_sheet") or "").strip()
            if not sheet_name:
                sheet_name = google_sheet_tabs[0] if google_sheet_tabs else "Order Sheet"
            items_by_sheet.setdefault(sheet_name, []).append(item)

        ordered_groups: list[tuple[str, list[dict]]] = []
        seen: set[str] = set()
        for title in google_sheet_tabs:
            ordered_groups.append((title, items_by_sheet.get(title, [])))
            seen.add(title)
        for title, sheet_items in items_by_sheet.items():
            if title not in seen:
                ordered_groups.append((title, sheet_items))
        return ordered_groups or [("Order Sheet", items)]

    def _count_unique_sheet_urls(self, items: list[dict], images_dir: str | None) -> int:
        groups = self._detect_product_groups(items)
        unique_urls: dict[str, bool] = {}
        for g in groups:
            url = g.get("image_url", "")
            if url:
                unique_urls[url] = True
        if images_dir:
            for item in items:
                url = item.get("approved_url", "")
                if url and url.startswith("http"):
                    unique_urls[url] = True
                for extra_url in item.get("additional_urls", []):
                    if extra_url and extra_url.startswith("http"):
                        unique_urls[extra_url] = True
        return len(unique_urls)

    def _make_sheet_name(self, title: str, used_names: set[str]) -> str:
        base = re.sub(r"[\[\]\*\:/\\\?]", "_", str(title or "Order Sheet")).strip() or "Order Sheet"
        base = base[:31] or "Order Sheet"
        candidate = base
        suffix = 2
        while candidate in used_names:
            extra = f" ({suffix})"
            candidate = f"{base[:max(1, 31 - len(extra))]}{extra}"
            suffix += 1
        used_names.add(candidate)
        return candidate

    def _populate_worksheet(
        self,
        *,
        wb: Workbook,
        ws,
        items: list[dict],
        images_dir: str | None,
        currency: str,
        progress_state: dict[str, int],
        tmp_images: list[str],
    ) -> None:
        # Build column list
        out_headers = list(STANDARD_COLUMNS.keys())
        if not _has_comming_soon_column(items):
            out_headers = [header for header in out_headers if header != "Comming Soon"]

        # ── Header row (row 2) ──
        SUMMARY_ROW = 1
        HEADER_ROW = 2
        DATA_START = 3

        has_grouped_cols = False
        for ci, header in enumerate(out_headers, 1):
            cell = ws.cell(row=HEADER_ROW, column=ci, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

            col_letter = get_column_letter(ci)
            cfg = STANDARD_COLUMNS.get(header, {"width": 14})
            w = cfg.get("width", 14)
            if w == 0:
                ws.column_dimensions[col_letter].outlineLevel = 1
                ws.column_dimensions[col_letter].hidden = True
                has_grouped_cols = True
            else:
                ws.column_dimensions[col_letter].width = w

        ws.row_dimensions[HEADER_ROW].height = 28

        if has_grouped_cols:
            ws.sheet_format.outlineLevelCol = 1

        # ── Detect product groups (items with same item_code get grouped) ──
        groups = self._detect_product_groups(items)
        row_to_group = {}
        for gi, g in enumerate(groups):
            for ri in range(g["start"], g["end"] + 1):
                row_to_group[ri] = (g, gi)

        # ── Download all images (parallel for speed) ──
        # Spill image bytes to a temp directory as soon as they arrive instead
        # of keeping every blob in RAM. Large exports (8k+ items, 1k+ unique
        # images) were OOM-killing the container during wb.save() because all
        # raw bytes lived in memory alongside the openpyxl workbook. The
        # downstream consumers (image embed + save_images_to_folder) read
        # from disk on demand.
        image_paths: dict[int, str] = {}      # group index → path on disk
        url_to_path: dict[str, str] = {}      # url → path on disk
        unique_urls = {}
        for gi, g in enumerate(groups):
            url = g.get("image_url", "")
            if url:
                unique_urls.setdefault(url, []).append(gi)

        if images_dir:
            for item in items:
                url = item.get("approved_url", "")
                if url and url.startswith(("http", "file://")) and url not in unique_urls:
                    unique_urls.setdefault(url, [])
                for extra_url in item.get("additional_urls", []):
                    if extra_url and extra_url.startswith(("http", "file://")) and extra_url not in unique_urls:
                        unique_urls.setdefault(extra_url, [])

        download_tmpdir = tempfile.mkdtemp(prefix="ordersheet_dl_")

        def _dl(url_gis):
            url, gis = url_gis
            return url, gis, self._download_image(url)

        with ThreadPoolExecutor(max_workers=16) as pool:
            for url, gis, img_bytes in pool.map(_dl, unique_urls.items()):
                progress_state["downloaded"] += 1
                self._report(progress_state["downloaded"], progress_state["total"], "downloading")
                if not img_bytes:
                    continue
                img_bytes.seek(0)
                raw = img_bytes.read()
                tmp_name = hashlib.md5(url.encode("utf-8")).hexdigest()[:24] + ".bin"
                tmp_path = os.path.join(download_tmpdir, tmp_name)
                try:
                    with open(tmp_path, "wb") as fh:
                        fh.write(raw)
                except OSError as exc:
                    logger.warning("image spill-to-disk failed (%s): %s", exc, url[:80])
                    del raw
                    continue
                del raw  # release the in-memory copy now that disk has it
                url_to_path[url] = tmp_path
                for gi in gis:
                    image_paths[gi] = tmp_path
        tmp_images.append(download_tmpdir)  # cleaned up after wb.save()

        # Column indices (1-based)
        col_idx = {h: i + 1 for i, h in enumerate(out_headers)}
        qty_col = col_idx.get("QTY")
        whs_col = col_idx.get("WHS Price")
        row_whs_col = col_idx.get("Row WHS Price")
        total_col = col_idx.get("QTY Total")
        pic_col = col_idx.get("Picture", 1)

        qty_letter = get_column_letter(qty_col) if qty_col else None
        whs_letter = get_column_letter(whs_col) if whs_col else None
        formula_whs_letter = get_column_letter(row_whs_col) if row_whs_col else whs_letter
        total_letter = get_column_letter(total_col) if total_col else None

        currency_symbol = detect_currency_symbol(items, currency)
        currency_fmt = f'"{currency_symbol}"#,##0.00'

        # ── Data rows (row 3+) ──
        for ri, item in enumerate(items):
            excel_row = ri + DATA_START
            g_info = row_to_group.get(ri)
            gi = g_info[1] if g_info else (ri % 2)
            g = g_info[0] if g_info else None

            num_in_group = g["end"] - g["start"] + 1 if g else 1
            ws.row_dimensions[excel_row].height = max(IMAGE_PT / num_in_group, TEXT_ROW_H) if g else self.row_height_pt

            gfill = GROUP_FILLS[gi % 2]
            pfill = PRICE_FILLS[gi % 2]

            for header in out_headers:
                ci = col_idx[header]
                cell = ws.cell(row=excel_row, column=ci)
                cell.border = THIN_BORDER

                if header == "Picture":
                    cell.fill = PICTURE_FILL
                    cell.alignment = CENTER

                elif header == "Brand Name":
                    cell.value = item.get("brand", "")
                    cell.fill = gfill
                    cell.font = BODY_FONT
                    cell.alignment = LEFT

                elif header == "Item Group":
                    cell.value = item.get("item_group", "")
                    cell.fill = gfill
                    cell.font = BODY_FONT
                    cell.alignment = LEFT

                elif header == "Manufacturer Code":
                    cell.value = item.get("item_code", "")
                    cell.fill = gfill
                    cell.font = BODY_FONT
                    cell.alignment = LEFT

                elif header == "Web Description 2":
                    cell.value = item.get("style_name", "")
                    cell.fill = gfill
                    cell.font = BODY_FONT
                    cell.alignment = LEFT

                elif header == "Barcode":
                    cell.value = item.get("barcode", "")
                    cell.fill = gfill
                    cell.font = BODY_FONT
                    cell.alignment = CENTER

                elif header == "Gender":
                    cell.value = item.get("gender", "")
                    cell.fill = gfill
                    cell.font = BODY_FONT
                    cell.alignment = CENTER

                elif header == "Color":
                    cell.value = item.get("color_name", "")
                    cell.fill = gfill
                    cell.font = BODY_FONT
                    cell.alignment = LEFT

                elif header == "Size":
                    cell.value = item.get("size", "")
                    cell.fill = gfill
                    cell.font = BODY_FONT
                    cell.alignment = CENTER

                elif header == "Stock":
                    qty = item.get("qty_available")
                    cell.value = float(qty) if qty else 0
                    cell.fill = FREESTOCK_FILL
                    cell.font = Font(size=9, color="166534", name="Calibri")
                    cell.alignment = CENTER
                    cell.number_format = "#,##0"

                elif header == "Comming Soon":
                    comming_soon = item.get("comming_soon_qty")
                    cell.value = _coerce_sheet_value(comming_soon)
                    cell.fill = gfill
                    cell.font = BODY_FONT
                    cell.alignment = CENTER

                elif header == "QTY":
                    cell.value = 0
                    cell.fill = QTY_FILL
                    cell.font = EDITABLE_FONT
                    cell.alignment = CENTER

                elif header == "QTY Total":
                    if qty_letter and formula_whs_letter:
                        cell.value = f"={qty_letter}{excel_row}*{formula_whs_letter}{excel_row}"
                    else:
                        cell.value = 0
                    cell.number_format = currency_fmt
                    cell.fill = TOTAL_FILL
                    cell.font = FORMULA_FONT
                    cell.alignment = RIGHT

                elif header == "WHS Price":
                    whs = item.get("wholesale_price")
                    cell.value = float(whs) if whs else None
                    cell.number_format = currency_fmt
                    cell.fill = pfill
                    cell.font = PRICE_FONT
                    cell.alignment = RIGHT

                elif header == "RRP Price":
                    rrp = item.get("retail_price")
                    cell.value = float(rrp) if rrp else None
                    cell.number_format = currency_fmt
                    cell.fill = pfill
                    cell.font = PRICE_FONT
                    cell.alignment = RIGHT

                elif header == "Pictures":
                    link_url = item.get("pictures_url", "") or item.get("approved_url", "")
                    if link_url and link_url.startswith("http"):
                        cell.value = "View"
                        cell.hyperlink = link_url
                        cell.font = LINK_FONT
                    else:
                        cell.value = ""
                        cell.font = BODY_FONT
                    cell.fill = gfill
                    cell.alignment = CENTER

                elif header == "ItemCode":
                    sap = item.get("sap_code", "")
                    if not sap:
                        ig = item.get("item_group", "")
                        sz = item.get("size", "")
                        sap = f"{ig} {sz}".strip() if ig else item.get("item_code", "")
                    cell.value = sap
                    cell.fill = gfill
                    cell.font = BODY_FONT
                    cell.alignment = LEFT

                elif header == "Row WHS Price":
                    whs = item.get("wholesale_price")
                    cell.value = float(whs) if whs else None
                    cell.number_format = currency_fmt
                elif header == "Row RRP Price":
                    rrp = item.get("retail_price")
                    cell.value = float(rrp) if rrp else None
                    cell.number_format = currency_fmt
                elif header == "Row Manufacturer Code":
                    cell.value = item.get("item_code", "")
                elif header == "Row Web Description 2":
                    cell.value = item.get("style_name", "")
                elif header == "Row Color":
                    cell.value = item.get("color_name", "")
                elif header == "Row Brand Name":
                    cell.value = item.get("brand", "")

            if images_dir:
                img_url = item.get("approved_url", "")
                if img_url and img_url in url_to_path:
                    self._save_image_file_from_path(url_to_path[img_url], images_dir, item)
                for extra_url in item.get("additional_urls", []):
                    if extra_url and extra_url in url_to_path:
                        self._save_image_file_from_path(url_to_path[extra_url], images_dir, item)

        last_data_row = max(DATA_START, DATA_START + len(items) - 1)

        # ── Merge cells for product groups & embed images ──
        for gi, g in enumerate(groups):
            excel_start = g["start"] + DATA_START
            excel_end = g["end"] + DATA_START

            if excel_end > excel_start:
                ws.merge_cells(
                    start_row=excel_start, start_column=pic_col,
                    end_row=excel_end, end_column=pic_col,
                )

            mc = ws.cell(row=excel_start, column=pic_col)
            mc.fill = PICTURE_FILL
            mc.alignment = CENTER
            mc.border = THIN_BORDER

            if gi not in image_paths:
                continue

            num_rows = g["end"] - g["start"] + 1
            row_h_pt = max(IMAGE_PT / num_rows, TEXT_ROW_H)
            total_cell_h_px = int(row_h_pt * num_rows / 0.75)
            display_h = min(IMAGE_PX, total_cell_h_px)

            # Open the spilled file directly — PIL streams from disk, so we
            # never hold the full-resolution bytes in RAM.
            try:
                pil_open = PILImage.open(image_paths[gi])
                pil_open.load()  # force decode now so the file handle can close
            except Exception as exc:
                logger.info("image load failed for gi=%s: %s", gi, exc)
                continue
            if pil_open.mode in ("RGBA", "LA", "P"):
                pil_rgba = pil_open.convert("RGBA")
                bg = PILImage.new("RGB", pil_rgba.size, (208, 208, 208))
                bg.paste(pil_rgba, mask=pil_rgba.split()[3])
                raw_img = bg
            else:
                raw_img = pil_open.convert("RGB")

            display_img = raw_img.copy()
            display_img.thumbnail((self.img_size[0], display_h), PILImage.LANCZOS)

            tmp = tempfile.NamedTemporaryFile(suffix=".jpg", delete=False)
            display_img.save(tmp.name, format="JPEG", quality=90, optimize=True)
            tmp.close()
            tmp_images.append(tmp.name)

            xl_img = XLImage(tmp.name)
            col_width_px = int(STANDARD_COLUMNS["Picture"]["width"] * 7 + 5)
            h_offset_emu = max(0, (col_width_px - display_img.width) // 2) * 9525
            v_offset_emu = max(0, (total_cell_h_px - display_img.height) // 2) * 9525

            try:
                anchor = OneCellAnchor()
                anchor._from = AnchorMarker(col=pic_col - 1, colOff=h_offset_emu,
                                            row=excel_start - 1, rowOff=v_offset_emu)
                anchor.ext = XDRPositiveSize2D(
                    cx=display_img.width * 9525,
                    cy=display_img.height * 9525,
                )
                xl_img.anchor = anchor
                ws.add_image(xl_img)
            except Exception:
                ws.add_image(xl_img, f"{get_column_letter(pic_col)}{excel_start}")

        # ── Summary row (row 1) ──
        for ci, header in enumerate(out_headers, 1):
            cell = ws.cell(row=SUMMARY_ROW, column=ci, value="")
            cell.fill = SUMMARY_FILL
            cell.border = THIN_BORDER

            if header == "QTY" and qty_letter:
                cell.value = f"=SUBTOTAL(109,{qty_letter}{DATA_START}:{qty_letter}{last_data_row})"
                cell.font = SUMMARY_QTY_FONT
                cell.alignment = CENTER
                cell.fill = QTY_FILL
            elif header == "QTY Total" and total_letter:
                cell.value = f"=SUBTOTAL(109,{total_letter}{DATA_START}:{total_letter}{last_data_row})"
                cell.number_format = currency_fmt
                cell.font = SUMMARY_TOTAL_FONT
                cell.alignment = RIGHT
                cell.fill = TOTAL_FILL
            else:
                cell.alignment = LEFT

        ws.row_dimensions[SUMMARY_ROW].height = 24

        # ── Freeze & filter ──
        ws.freeze_panes = f"B{DATA_START}"
        last_col_letter = get_column_letter(len(out_headers))
        filter_last_row = max(last_data_row, HEADER_ROW)
        ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col_letter}{filter_last_row}"

        ws.sheet_properties.tabColor = "1F2937"
        ws.print_title_rows = "1:2"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1

    def _detect_product_groups(self, items: list[dict]) -> list[dict]:
        """Group items by item_code (same product, different sizes/colors)."""
        groups = []
        if not items:
            return groups

        current_code = items[0].get("item_code", "")
        start = 0

        for i in range(1, len(items)):
            code = items[i].get("item_code", "")
            if code != current_code:
                groups.append({
                    "start": start,
                    "end": i - 1,
                    "image_url": _best_image_url_for_group(items, start, i - 1),
                    "group_index": len(groups),
                })
                start = i
                current_code = code

        # Last group
        groups.append({
            "start": start,
            "end": len(items) - 1,
            "image_url": _best_image_url_for_group(items, start, len(items) - 1),
            "group_index": len(groups),
        })

        return groups

    def _download_image(self, url: str) -> io.BytesIO | None:
        """Download image at full resolution (no thumbnail)."""
        if not url:
            return None

        # Local file uploaded to server — read directly from disk
        if url.startswith("file://"):
            path = url[7:]  # strip "file://" → absolute server path
            try:
                with open(path, "rb") as f:
                    buf = io.BytesIO(f.read())
                PILImage.open(buf).verify()
                buf.seek(0)
                return buf
            except Exception:
                return None

        if not url.startswith("http"):
            return None

        # Dropbox sharing URLs need to be turned into direct downloads, otherwise
        # we get an HTML preview page instead of the image bytes.
        dl_url = url
        if "dropbox.com" in dl_url or "dropboxusercontent.com" in dl_url:
            if "dl=0" in dl_url:
                dl_url = dl_url.replace("dl=0", "dl=1")
            elif "dl=1" not in dl_url:
                sep = "&" if "?" in dl_url else "?"
                dl_url = f"{dl_url}{sep}dl=1"

        try:
            # Stream the download so a giant image (Dropbox originals can be
            # 30 MB each) can't single-handedly blow up memory before we
            # check the Content-Length.
            with requests.get(dl_url, headers=_DL_HEADERS, timeout=20, stream=True) as resp:
                if resp.status_code != 200:
                    logger.info("image download skipped (status %s): %s",
                                resp.status_code, url[:120])
                    return None
                _MAX_BYTES = 20_000_000  # 20 MB hard cap per image
                chunks: list[bytes] = []
                seen = 0
                for chunk in resp.iter_content(chunk_size=131072):
                    if not chunk:
                        continue
                    seen += len(chunk)
                    if seen > _MAX_BYTES:
                        logger.info("image too large (>%d bytes), truncating: %s",
                                    _MAX_BYTES, url[:120])
                        chunks.append(chunk[:_MAX_BYTES - (seen - len(chunk))])
                        break
                    chunks.append(chunk)
                raw = b"".join(chunks)
            buf = io.BytesIO(raw)
            # Validate it's actually an image
            PILImage.open(buf).verify()
            # Immediately downsample very large images so we don't carry
            # 30 MB blobs for every URL until wb.save() runs. The embed
            # step thumbnails again to its display size; the on-disk copy
            # gets downsampled to 1600 px in _save_image_file.
            buf.seek(0)
            if len(raw) > 3_000_000:
                try:
                    from PIL import ImageOps  # noqa: PLC0415
                    img = PILImage.open(buf)
                    img = ImageOps.exif_transpose(img)
                    if img.mode != "RGB":
                        img = img.convert("RGB")
                    if max(img.size) > 1600:
                        img.thumbnail((1600, 1600), PILImage.LANCZOS)
                    small = io.BytesIO()
                    img.save(small, format="JPEG", quality=88, optimize=True)
                    small.seek(0)
                    return small
                except Exception:
                    pass  # fall through to returning original bytes
            buf.seek(0)
            return buf
        except Exception as exc:
            logger.info("image download failed (%s): %s", exc, url[:120])
            return None

    def _save_image_file_from_path(self, src_path: str, base_images_dir: str, item: dict):
        """Disk-friendly variant of `_save_image_file` — reads bytes from
        `src_path` instead of taking them as an argument. Cuts peak RAM
        because the caller no longer has to hold every image's bytes alive
        until the end of the worksheet loop.
        """
        try:
            with open(src_path, "rb") as fh:
                data = fh.read()
        except OSError as exc:
            logger.info("save_image: cannot read %s (%s)", src_path, exc)
            return
        self._save_image_file(data, base_images_dir, item)

    def _save_image_file(self, img_data: bytes, base_images_dir: str, item: dict):
        """
        Save full-resolution image to folder.

        Naming convention for the B2B importer:
          • Folder name  = Item Group Code, with underscores converted to spaces
          • Highlight    = {item_code}_01.jpg
          • Additional   = {item_code}_02.jpg, _03.jpg, ...

        The first image saved for an item is always _01 so the customer-
        facing highlight image is deterministic regardless of colour code or
        candidate ordering.
        """
        item_code = str(item.get("item_code") or "unknown").strip() or "unknown"
        folder_code = str(
            item.get("item_group_code") or item.get("sap_code") or item.get("item_group") or ""
        ).strip()
        safe_code = re.sub(r"[^\w\-]", "_", item_code)

        # Folder names follow Excel/Google "Item Group Code" first, then
        # SAP ItemCode, and use single spaces. Image file names keep
        # underscores: ITEMCODE_01.jpg.
        folder_name = normalize_folder_name(folder_code, default=safe_code)

        folder_path = os.path.join(base_images_dir, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        # The B2B importer accepts JPG/PNG reliably. Keep those formats as-is,
        # but convert WebP/GIF/TIFF/BMP/unknown images to JPG.
        ext = ".jpg"
        if img_data[:8].startswith(b"\x89PNG\r\n\x1a\n"):
            ext = ".png"
        elif img_data[:2] == b"\xff\xd8":
            ext = ".jpg"

        # Find next available 2-digit number across any image extension; the
        # primary image is always _01 regardless of suffix.
        n = 1
        while any(
            os.path.exists(os.path.join(folder_path, f"{safe_code}_{n:02d}{x}"))
            for x in (".jpg", ".jpeg", ".png")
        ):
            n += 1

        path = os.path.join(folder_path, f"{safe_code}_{n:02d}{ext}")

        # Path 1 — JPG/PNG that's already a reasonable size: write source bytes
        # verbatim (no PIL round-trip means no colour shifts and no
        # transparency artifacts). For oversized originals (Dropbox masters
        # are often 5–30 MB each, which adds up to tens of GB across an
        # 8000-item export and blew up the disk twice) we fall through to
        # the PIL path below to downsample first.
        _MAX_RAW_BYTES = 1_500_000  # ~1.5 MB
        if ext in (".jpg", ".png") and img_data[:2] in (b"\xff\xd8", b"\x89P") \
                and len(img_data) <= _MAX_RAW_BYTES:
            try:
                with open(path, "wb") as f:
                    f.write(img_data)
                return
            except OSError:
                pass

        # Path 2 — exotic / corrupt header OR an oversized JPG/PNG: decode
        # with PIL, downsample to a reasonable max dimension, composite
        # transparency onto white, save as JPEG. This is what keeps the
        # output volume from filling up on big sessions.
        try:
            from PIL import ImageOps  # noqa: PLC0415
            img = PILImage.open(io.BytesIO(img_data))
            img = ImageOps.exif_transpose(img)
            if img.mode in ("RGBA", "LA"):
                bg = PILImage.new("RGB", img.size, (255, 255, 255))
                bg.paste(img, mask=img.split()[-1])
                img = bg
            elif img.mode == "P":
                img = img.convert("RGBA")
                bg = PILImage.new("RGB", img.size, (255, 255, 255))
                bg.paste(img, mask=img.split()[-1] if "A" in img.getbands() else None)
                img = bg
            elif img.mode != "RGB":
                img = img.convert("RGB")
            # Long-edge cap of 1600 px is plenty for the B2B importer; an
            # 8000-item session at this resolution stays under ~5 GB instead
            # of blowing past 30 GB of multi-MB Dropbox originals.
            _MAX_DIM = 1600
            if max(img.size) > _MAX_DIM:
                img.thumbnail((_MAX_DIM, _MAX_DIM), PILImage.LANCZOS)
            jpg_path = os.path.join(folder_path, f"{safe_code}_{n:02d}.jpg")
            img.save(jpg_path, format="JPEG", quality=88, optimize=True)
        except Exception:
            # Last-ditch fallback: write raw bytes. Only happens if PIL
            # couldn't decode the image at all.
            with open(path, "wb") as f:
                f.write(img_data)
