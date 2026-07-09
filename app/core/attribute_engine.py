"""Product attribute engine — read a SAP product export, classify each style's
SAP product type + FABRIC/FIT/STYLE/WEIGHT (constrained to SAP's value lists),
and build the SAP attribute upload sheet (Style Code | Item Group | Attribute | Value).
"""
from __future__ import annotations

import json
import logging
import re

import openpyxl

from app.core.attribute_taxonomy import (
    PRODUCT_TYPES_BY_GROUP,
    VALUE_LISTS,
    master_for_item_group,
)
from app.services.ai_service import _call_ai

logger = logging.getLogger(__name__)


def _norm(s) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


# Candidate header names (normalized) for each field we need. Order matters:
# the first alias found wins ("SAP Item Group" must beat "Item Group Code",
# which in Carhartt exports holds the full item code, not the group).
_HEADER_ALIASES = {
    "style_code": ["stylecode", "mfrcatalogno", "manufacturercode", "code"],
    "item_group": ["itemgroup", "sapitemgroup", "itemgroupcode"],
    "name": ["webdescription2", "itemdescriptionlong", "name", "description"],
    "material": ["material"],
    "gender": ["gender"],
    "vendor_category": ["vendorcategory"],
    "long_description": ["longdescription"],
    "season": ["season"],
}

# Vendor long descriptions arrive as one field with µ-separated catalog
# attributes ("8 x 8 cm µ 30 Pack µ screen print") — split for readability.
_LONG_DESC_MAX = 600


def _clean_long_description(s: str) -> str:
    s = re.sub(r"\s*µ\s*", "; ", str(s or ""))
    return re.sub(r"\s+", " ", s).strip()[:_LONG_DESC_MAX]


def parse_sap_products(path: str) -> tuple[list[dict], dict]:
    """Parse a SAP product export. Returns (styles, meta).

    One entry per distinct style code (deduped across size/colour rows).
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    # Find the header row (first row in the top 5 that has our key columns).
    header = None
    header_idx = 0
    buffered = []
    for i, row in enumerate(rows):
        buffered.append(row)
        norm = [_norm(c) for c in row]
        if "stylecode" in norm or ("itemgroup" in norm and any(
                a in norm for a in _HEADER_ALIASES["name"])):
            header = row
            header_idx = i
            break
        if i >= 6:
            break
    if header is None:
        header = buffered[0] if buffered else []

    col = {}
    norm_header = [_norm(c) for c in header]
    for field, aliases in _HEADER_ALIASES.items():
        for a in aliases:
            if a in norm_header:
                col[field] = norm_header.index(a)
                break

    if "style_code" not in col:
        raise ValueError(
            "Could not find a 'Style Code' column (or Mfr. Catalog No. / Code). "
            "Make sure this is a SAP product export."
        )

    def get(row, field):
        idx = col.get(field)
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    styles: dict[str, dict] = {}
    # Continue iterating the remaining rows after the header.
    for row in rows:
        sc = get(row, "style_code")
        if not sc:
            continue
        if sc not in styles:
            ig = get(row, "item_group")
            styles[sc] = {
                "style_code": sc,
                "name": get(row, "name"),
                "item_group": ig,
                "master_group": master_for_item_group(ig),
                "material": get(row, "material"),
                "gender": get(row, "gender"),
                "vendor_category": get(row, "vendor_category"),
                "long_description": _clean_long_description(get(row, "long_description")),
                "season": get(row, "season"),
            }
    wb.close()
    meta = {"columns_found": list(col.keys()), "row_count": header_idx}
    return list(styles.values()), meta


def _parse_json(text):
    if not text:
        return None
    text = re.sub(r"^```(?:json)?", "", text.strip()).strip()
    text = re.sub(r"```$", "", text).strip()
    m = re.search(r"\{.*\}", text, re.DOTALL)
    try:
        return json.loads(m.group(0)) if m else None
    except Exception:
        return None


def enrich_style(style: dict) -> dict:
    """One AI call: SAP product type + FABRIC/FIT/STYLE/WEIGHT, validated."""
    mg = style["master_group"]
    cands = PRODUCT_TYPES_BY_GROUP.get(mg, [])
    cand_txt = ", ".join(f"{c}={n}" for c, n in cands)
    valid_pt = {c for c, _ in cands}
    product = (
        f"PRODUCT: name={style.get('name')}; material={style.get('material')}; "
        f"group={style.get('item_group')}; gender={style.get('gender')}"
    )
    if style.get("vendor_category"):
        product += f"; vendor_category={style['vendor_category']}"
    if style.get("season"):
        product += f"; season={style['season']}"
    extra = ""
    if style.get("long_description"):
        extra += f"\nVENDOR LONG DESCRIPTION: {style['long_description']}"
    if style.get("ref_text"):
        extra += f"\nCATALOG/LOOKBOOK EXTRACT (matched by style number): {style['ref_text']}"
    prompt = (
        "Assign SAP attributes to a fashion product. Use ONLY the allowed values "
        "(exact codes). Be conservative when the description is thin. When a "
        "vendor long description or catalog extract is provided, trust it — it "
        "carries the real fabric/fit/style detail.\n"
        f"1) product_type: exactly ONE of [{cand_txt}]\n"
        f"2) FABRIC: one of {VALUE_LISTS['FABRIC']} or null "
        "(special fabrics only; basic cotton = null)\n"
        f"3) FIT: one of {VALUE_LISTS['FIT']} or null\n"
        f"4) STYLE: up to 2 of {VALUE_LISTS['STYLE']} or []\n"
        f"5) WEIGHT: one of {VALUE_LISTS['WEIGHT']} or null\n"
        'Return ONLY JSON: {"product_type":"CODE","confidence":0.0,'
        '"FABRIC":null,"FIT":null,"STYLE":[],"WEIGHT":null}\n\n'
        + product + extra
    )
    data = {}
    for _ in range(3):
        data = _parse_json(_call_ai(prompt, max_tokens=300)) or {}
        if data.get("product_type") in valid_pt:
            break

    pt = data.get("product_type") if data.get("product_type") in valid_pt else None
    fab = data.get("FABRIC") if data.get("FABRIC") in VALUE_LISTS["FABRIC"] else None
    fit = data.get("FIT") if data.get("FIT") in VALUE_LISTS["FIT"] else None
    wt = data.get("WEIGHT") if data.get("WEIGHT") in VALUE_LISTS["WEIGHT"] else None
    style_vals = [s for s in (data.get("STYLE") or []) if s in VALUE_LISTS["STYLE"]][:2]
    try:
        conf = float(data.get("confidence"))
    except (TypeError, ValueError):
        conf = 0.0
    needs_review = (pt is None) or (conf < 0.5)
    return {
        **style,
        "product_type": pt,
        "confidence": conf,
        "FABRIC": fab, "FIT": fit, "WEIGHT": wt, "STYLE": style_vals,
        "needs_review": needs_review,
    }


def build_upload_workbook(results: list[dict], path: str) -> dict:
    """Write the SAP upload xlsx (clean rows only). Returns summary counts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Style Code", "Item Group", "Attribute Code", "Value"])
    rows = 0
    clean = 0
    for r in results:
        if r.get("needs_review") or not r.get("product_type"):
            continue
        clean += 1
        sc, grp = r["style_code"], r["master_group"]
        ws.append([sc, grp, r["product_type"], "Y"]); rows += 1
        if r.get("FABRIC"):
            ws.append([sc, grp, "FABRIC", r["FABRIC"]]); rows += 1
        if r.get("FIT"):
            ws.append([sc, grp, "FIT", r["FIT"]]); rows += 1
        if r.get("WEIGHT"):
            ws.append([sc, grp, "WEIGHT", r["WEIGHT"]]); rows += 1
        for s in r.get("STYLE", []):
            ws.append([sc, grp, "STYLE", s]); rows += 1
    wb.save(path)
    return {"clean_styles": clean, "rows": rows,
            "review_styles": sum(1 for r in results if r.get("needs_review"))}
