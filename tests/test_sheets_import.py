from __future__ import annotations

import io
import importlib
import time

from openpyxl import load_workbook

from app.core.sheets_reader import SheetsReader, is_preorder_format


# Header row of the SAP Preorder/Reorder order-document layout (subset, same
# column order as the live sheet). One tab per order doc, one line per row.
_PREORDER_HEADERS = [
    "Picture", "Pictures", "Type", "DocNum", "Brand", "ITEM CODE",
    "DESCRIPTION", "Category", "GENDER", "Size", "Color", "BARCODE",
    "QTY", "Unit Price", "Currency", "SRP price", "SAP CODE",
]


def _img(url):
    return f'=IMAGE("{url}")'


def _link(url, label):
    return f'=HYPERLINK("{url}","{label}")'


def test_is_preorder_format_distinguishes_layouts():
    assert is_preorder_format(_PREORDER_HEADERS) is True
    # Reorder tabs share the identical schema
    assert is_preorder_format(["Picture", "Pictures", "Type", "DocNum",
                               "ITEM CODE", "QTY", "Unit Price"]) is True
    # Stock Ordersheet layout must NOT be misdetected
    assert is_preorder_format([
        "Picture", "Manufacturer Code", "Brand Name", "Color", "Size",
        "WHS Price", "RRP Price", "Stock",
    ]) is False
    # Invoice tabs (no DocNum/Unit Price) must NOT match the preorder extractor
    assert is_preorder_format([
        "Picture", "Pictures", "Invoice No", "ITEM CODE", "DESCRIPTION",
    ]) is False


def test_preorder_extractor_maps_columns_and_carries_image_forward():
    a_img = "https://dl.dropboxusercontent.com/scl/fi/aaa/tee.png?rlkey=x"
    b_img = "https://dl.dropboxusercontent.com/scl/fi/bbb/cap.png?rlkey=y"
    a_folder = "https://www.dropbox.com/scl/fo/folderA"
    b_folder = "https://www.dropbox.com/scl/fo/folderB"

    # Product A spans two size rows; the =IMAGE photo lives on the first row only.
    # Product B is a single row. Trailing blank row must be ignored.
    display_rows = [
        ["", "ACODE BLACK", "Preorder", "3034", "TestBrand", "ACODE BLACK",
         "Test Tee", "T-Shirts", "Men", "M", "BLACK", "BAR-M",
         "2", "21.03", "USD", "50.00 USD", "SAP-A-M"],
        ["", "", "Preorder", "3034", "TestBrand", "ACODE BLACK",
         "Test Tee", "T-Shirts", "Men", "L", "BLACK", "BAR-L",
         "1", "21.03", "USD", "50.00 USD", "SAP-A-L"],
        ["", "BCODE NAVY", "Preorder", "3034", "TestBrand", "BCODE NAVY",
         "Test Cap", "Headwear", "Men", "onesize", "NAVY", "BAR-OS",
         "3", "30.00", "USD", "70.00 USD", "SAP-B"],
        ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    formula_rows = [
        [_img(a_img), _link(a_folder, "ACODE BLACK")] + display_rows[0][2:],
        ["", ""] + display_rows[1][2:],
        [_img(b_img), _link(b_folder, "BCODE NAVY")] + display_rows[2][2:],
        display_rows[3],
    ]
    tab = {
        "title": "Preorder3034",
        "headers": _PREORDER_HEADERS,
        "display_rows": display_rows,
        "formula_rows": formula_rows,
    }

    reader = SheetsReader.__new__(SheetsReader)  # skip credential loading
    items = reader.extract_items_from_tab(tab)

    assert len(items) == 3  # trailing blank row dropped

    first, second, third = items
    assert first["item_code"] == "ACODE BLACK"
    assert first["brand"] == "TestBrand"
    assert first["style_name"] == "Test Tee"
    assert first["item_group"] == "T-Shirts"
    assert first["color_name"] == "BLACK"
    assert first["size"] == "M"
    assert first["gender"] == "Men"
    assert first["barcode"] == "BAR-M"
    assert first["wholesale_price"] == "21.03"
    assert first["retail_price"] == "50.00 USD"
    assert first["ordered_qty"] == "2"
    assert first["qty_available"] == "2"
    assert first["currency"] == "USD"
    assert first["sap_code"] == "SAP-A-M"
    assert first["image_url"] == a_img
    assert a_folder in first["dropbox_url"]

    # Second size row has no photo of its own — it inherits product A's image.
    assert second["size"] == "L"
    assert second["image_url"] == a_img
    assert a_folder in second["dropbox_url"]

    # New product resets the carried image.
    assert third["item_code"] == "BCODE NAVY"
    assert third["image_url"] == b_img
    assert b_folder in third["dropbox_url"]



def test_expand_batch_jobs_keeps_one_job_per_sheet_url(test_app):
    sheets_routes = test_app["sheets_routes"]
    url_one = "https://docs.google.com/spreadsheets/d/abc123/edit#gid=0"
    url_two = "https://docs.google.com/spreadsheets/d/xyz789/edit#gid=0"

    jobs = sheets_routes._expand_batch_jobs(
        [url_one, url_two],
        {
            url_one: ["ReOrder_Dubai", "PreOrder_Carhartt WIP_2026-04"],
            url_two: ["Main"],
        },
    )

    assert len(jobs) == 2
    assert jobs[0]["url"] == url_one
    assert jobs[0]["selected_tabs"] == ["ReOrder_Dubai", "PreOrder_Carhartt WIP_2026-04"]
    assert jobs[1]["url"] == url_two
    assert jobs[1]["selected_tabs"] == ["Main"]


def test_import_batch_initializes_one_parallel_job_per_sheet_url(
    client,
    login_as,
    test_app,
    monkeypatch,
):
    login_as()
    sheets_routes = test_app["sheets_routes"]

    cred_path = test_app["temp_root"] / "google-test-creds.json"
    cred_path.write_text("{}", encoding="utf-8")
    monkeypatch.setattr(sheets_routes, "_get_credentials_path", lambda _uid: str(cred_path))

    def fake_create_task(coro):
        coro.close()

        class _DummyTask:
            pass

        return _DummyTask()

    monkeypatch.setattr(sheets_routes.asyncio, "create_task", fake_create_task)

    url_one = "https://docs.google.com/spreadsheets/d/abc123/edit#gid=0"
    url_two = "https://docs.google.com/spreadsheets/d/xyz789/edit#gid=0"

    response = client.post(
        "/sheets/import-batch",
        json={
            "urls": [url_one, url_two],
            "selected_tabs": {
                url_one: ["Sheet One", "Sheet Two"],
                url_two: ["Main"],
            },
            "save_images": True,
            "search_missing": True,
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["ok"] is True
    assert data["total"] == 2

    batch = sheets_routes._batch_progress[data["batch_id"]]
    assert batch["total"] == 2
    assert batch["jobs"][0]["url"] == url_one
    assert batch["jobs"][1]["url"] == url_two


def test_google_sheet_import_persists_selected_tab_order_and_source_sheet(
    make_user,
    test_app,
    monkeypatch,
):
    user = make_user()
    sheets_routes = test_app["sheets_routes"]
    sheets_reader = importlib.import_module("app.core.sheets_reader")
    models = test_app["models"]

    class FakeSheetsReader:
        def __init__(self, _cred_path):
            pass

        def fetch_spreadsheet(self, _spreadsheet_id):
            return {
                "title": "Buying Sheet",
                "tabs": [
                    {"title": "Tab A", "headers": ["WHS Price"]},
                    {"title": "Tab B", "headers": ["WHS Price"]},
                    {"title": "Tab C", "headers": ["WHS Price"]},
                ],
            }

        def extract_items_from_tab(self, tab):
            title = tab["title"]
            return [{
                "item_code": f"SKU-{title[-1]}",
                "size": "42",
                "brand": "Brand",
                "style_name": f"Style {title[-1]}",
                "color_name": "Black",
                "gender": "Men",
                "wholesale_price": "10",
                "retail_price": "20",
                "qty_available": "5",
                "barcode": f"BAR-{title[-1]}",
                "item_group": "Shoes",
                "item_group_code": f"GRP-{title[-1]}",
                "sap_code": f"SAP-{title[-1]}",
                "image_url": "",
                "dropbox_url": "",
                "comming_soon_qty": "",
            }]

    monkeypatch.setattr(sheets_reader, "SheetsReader", FakeSheetsReader)
    monkeypatch.setattr(sheets_reader, "extract_spreadsheet_id", lambda _url: "sheet-123")

    result = sheets_routes._do_import_sheet_sync(
        user["id"],
        "https://docs.google.com/spreadsheets/d/sheet-123/edit#gid=0",
        "unused-creds.json",
        selected_tabs=["Tab B", "Tab A"],
    )

    assert result["ok"] is True

    db = test_app["database"].SessionLocal()
    try:
        sess = db.get(models.Session, result["session_id"])
        items = (
            db.query(models.UniqueItem)
            .filter(models.UniqueItem.session_id == sess.id)
            .order_by(models.UniqueItem.id.asc())
            .all()
        )
        assert sess.config["selected_sheet_tabs"] == ["Tab A", "Tab B"]
        assert [item.source_sheet for item in items] == ["Tab A", "Tab B"]
        assert [item.item_group_code for item in items] == ["GRP-A", "GRP-B"]
        assert [item.sap_code for item in items] == ["SAP-A", "SAP-B"]
    finally:
        db.close()


def test_google_sheet_import_deduplicates_across_tabs_by_source_sheet(
    make_user,
    test_app,
    monkeypatch,
):
    """Same SKU+color+size on two different tabs must both be persisted."""
    user = make_user()
    sheets_routes = test_app["sheets_routes"]
    sheets_reader = importlib.import_module("app.core.sheets_reader")
    models = test_app["models"]

    class FakeSheetsReader:
        def __init__(self, _cred_path):
            pass

        def fetch_spreadsheet(self, _spreadsheet_id):
            return {
                "title": "Buying Sheet",
                "tabs": [
                    {"title": "Spring", "headers": ["WHS Price"]},
                    {"title": "Summer", "headers": ["WHS Price"]},
                ],
            }

        def extract_items_from_tab(self, tab):
            # SAME item_code + color_name + size on both tabs
            return [{
                "item_code": "SKU-001",
                "size": "42",
                "brand": "Brand",
                "style_name": "Style X",
                "color_name": "Black",
                "gender": "Men",
                "wholesale_price": "10",
                "retail_price": "20",
                "qty_available": "5",
                "barcode": "BAR-001",
                "item_group": "Shoes",
                "item_group_code": "GRP-001",
                "sap_code": "SAP-001",
                "image_url": "",
                "dropbox_url": "",
                "comming_soon_qty": "",
            }]

    monkeypatch.setattr(sheets_reader, "SheetsReader", FakeSheetsReader)
    monkeypatch.setattr(sheets_reader, "extract_spreadsheet_id", lambda _url: "sheet-456")

    result = sheets_routes._do_import_sheet_sync(
        user["id"],
        "https://docs.google.com/spreadsheets/d/sheet-456/edit#gid=0",
        "unused-creds.json",
        selected_tabs=["Spring", "Summer"],
    )

    assert result["ok"] is True
    assert result["items"] == 2, "both tabs' items must survive despite identical SKU+color+size"

    db = test_app["database"].SessionLocal()
    try:
        items = (
            db.query(models.UniqueItem)
            .filter(models.UniqueItem.session_id == result["session_id"])
            .order_by(models.UniqueItem.source_sheet.asc())
            .all()
        )
        assert len(items) == 2
        source_sheets = {item.source_sheet for item in items}
        assert source_sheets == {"Spring", "Summer"}
        assert {item.item_group_code for item in items} == {"GRP-001"}
        assert {item.sap_code for item in items} == {"SAP-001"}
        # color_codes must differ so the unique constraint isn't violated
        assert items[0].color_code != items[1].color_code
    finally:
        db.close()


def test_order_mode_export_fills_qty_and_drops_stock(tmp_path):
    """Preorder/Reorder exports reproduce the placed order: QTY holds the ordered
    quantity, line totals compute from it, and the 'Stock' column is dropped."""
    from openpyxl import load_workbook

    from app.core.generator import OrderSheetGenerator

    items = [
        {
            "item_code": "ACODE BLACK", "brand": "TestBrand", "style_name": "Tee",
            "item_group": "T-Shirts", "color_name": "BLACK", "size": "M",
            "gender": "Men", "barcode": "BAR-M",
            "wholesale_price": 21.03, "retail_price": 50.0,
            "qty_available": 2, "source_sheet": "Preorder3034",
        },
        {
            "item_code": "ACODE BLACK", "brand": "TestBrand", "style_name": "Tee",
            "item_group": "T-Shirts", "color_name": "BLACK", "size": "L",
            "gender": "Men", "barcode": "BAR-L",
            "wholesale_price": 21.03, "retail_price": 50.0,
            "qty_available": 1, "source_sheet": "Preorder3034",
        },
    ]

    gen = OrderSheetGenerator(config={"image_size": [150, 150], "row_height_px": 100})
    path = gen.generate(items, str(tmp_path), input_filename="Preorder3034",
                        brand="TestBrand", currency="$",
                        google_sheet_tabs=["Preorder3034"], order_mode=True)

    wb = load_workbook(path)
    try:
        ws = wb[wb.sheetnames[0]]
        headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
        assert "Stock" not in headers           # free stock is meaningless for an order
        qty_col = headers.index("QTY") + 1
        total_col = headers.index("QTY Total") + 1
        # QTY shows the actual ordered quantities (not a blank 0 buying form)
        assert ws.cell(3, qty_col).value == 2
        assert ws.cell(4, qty_col).value == 1
        # Line total is a formula multiplying QTY by the (hidden) row WHS price
        assert str(ws.cell(3, total_col).value).startswith("=")
    finally:
        wb.close()


def test_backfill_sap_code_repairs_existing_google_sheet_sessions(
    make_user,
    test_app,
    monkeypatch,
    tmp_path,
):
    user = make_user()
    models = test_app["models"]
    backfill = importlib.import_module("app.services.sap_code_backfill")

    class FakeSheetsReader:
        def __init__(self, _cred_path):
            pass

        def fetch_spreadsheet(self, _spreadsheet_id):
            return {"tabs": [{"title": "ReOrder_Dubai"}]}

        def extract_items_from_tab(self, _tab):
            return [{
                "item_code": "ACL-253-SC-447-001",
                "size": "M",
                "color_name": "Black",
                "item_group_code": "ACL_A_ACL-253-SC-447-001_Black",
                "sap_code": "ACL253SC447001",
            }]

    cred_path = tmp_path / "google.json"
    cred_path.write_text("{}", encoding="utf-8")
    monkeypatch.setattr(backfill, "_credentials_path", lambda _uid: str(cred_path))
    monkeypatch.setattr(backfill, "SheetsReader", FakeSheetsReader)
    monkeypatch.setattr(backfill, "extract_spreadsheet_id", lambda _url: "sheet-123")

    db = test_app["database"].SessionLocal()
    try:
        sess = models.Session(
            user_id=user["id"],
            name="Existing Google Session",
            source_type="google_sheets",
            source_ref="https://docs.google.com/spreadsheets/d/sheet-123/edit",
            status="reviewing",
        )
        sess.config = {"selected_sheet_tabs": ["ReOrder_Dubai"]}
        db.add(sess)
        db.commit()
        db.refresh(sess)

        item = models.UniqueItem(
            session_id=sess.id,
            item_code="ACL-253-SC-447-001",
            color_name="Black",
            color_code="Black|M|ReOrder_Dubai",
            item_group="ACCS",
            item_group_code="",
            source_sheet="ReOrder_Dubai",
            review_status="approved",
            search_status="done",
        )
        item.sizes = ["M"]
        db.add(item)
        db.commit()

        updated = backfill.backfill_sap_codes_for_session(db, sess, user["id"])
        db.refresh(item)
        assert updated == 1
        assert item.item_group_code == "ACL_A_ACL-253-SC-447-001_Black"
        assert item.sap_code == "ACL253SC447001"
    finally:
        db.close()


def test_google_sheet_export_keeps_selected_tabs_as_workbook_sheets(
    client,
    login_as,
    test_app,
):
    user = login_as()
    models = test_app["models"]
    db = test_app["database"].SessionLocal()
    try:
        sess = models.Session(
            user_id=user["id"],
            name="Google Multi Tab",
            source_type="google_sheets",
            source_ref="https://docs.google.com/spreadsheets/d/sheet-123/edit#gid=0",
            status="reviewing",
            total_items=2,
            searched_items=2,
        )
        sess.config = {
            "selected_sheet_tabs": ["Tab A", "Tab B", "Tab C"],
            "currency": "€",
        }
        db.add(sess)
        db.commit()
        db.refresh(sess)

        db.add_all([
            models.UniqueItem(
                session_id=sess.id,
                item_code="SKU-A",
                color_code="Black|42",
                brand="Brand",
                style_name="Style A",
                color_name="Black",
                gender="Men",
                wholesale_price=10,
                retail_price=20,
                qty_available=5,
                review_status="approved",
                source_sheet="Tab A",
                sizes=["42"],
            ),
            models.UniqueItem(
                session_id=sess.id,
                item_code="SKU-C",
                color_code="Brown|43",
                brand="Brand",
                style_name="Style C",
                color_name="Brown",
                gender="Men",
                wholesale_price=11,
                retail_price=21,
                qty_available=6,
                review_status="approved",
                source_sheet="Tab C",
                sizes=["43"],
            ),
        ])
        db.commit()
        session_id = sess.id
    finally:
        db.close()

    response = client.post(f"/generate/{session_id}", json={"save_images": False})
    assert response.status_code == 200
    assert response.json()["ok"] is True

    excel_file = None
    for _ in range(40):
        poll_db = test_app["database"].SessionLocal()
        try:
            excel_file = (
                poll_db.query(models.GeneratedFile)
                .filter(
                    models.GeneratedFile.session_id == session_id,
                    models.GeneratedFile.filename != "images.zip",
                )
                .first()
            )
            if excel_file is not None:
                poll_db.expunge(excel_file)
                break
        finally:
            poll_db.close()
        time.sleep(0.25)

    assert excel_file is not None

    download = client.get(f"/download/{excel_file.token}")
    assert download.status_code == 200

    wb = load_workbook(io.BytesIO(download.content))
    try:
        assert wb.sheetnames == ["Tab A", "Tab B", "Tab C"]
        assert wb["Tab A"]["D3"].value == "SKU-A"
        assert wb["Tab C"]["D3"].value == "SKU-C"
        assert wb["Tab B"]["A2"].value == "Picture"
        assert wb["Tab B"].max_row == 2
    finally:
        wb.close()


def test_google_sheet_convert_only_session_exports_without_image_review(
    client,
    login_as,
    test_app,
):
    user = login_as()
    models = test_app["models"]
    db = test_app["database"].SessionLocal()
    try:
        sess = models.Session(
            user_id=user["id"],
            name="Convert Only",
            source_type="google_sheets",
            source_ref="https://docs.google.com/spreadsheets/d/sheet-456/edit#gid=0",
            status="reviewing",
            total_items=1,
            searched_items=1,
        )
        sess.config = {
            "selected_sheet_tabs": ["Main"],
            "currency": "€",
            "search_missing": False,
        }
        db.add(sess)
        db.commit()
        db.refresh(sess)

        db.add(
            models.UniqueItem(
                session_id=sess.id,
                item_code="SKU-CONVERT",
                color_code="Black|42|Main",
                brand="Brand",
                style_name="Converter Shoe",
                color_name="Black",
                gender="Men",
                wholesale_price=10,
                retail_price=20,
                qty_available=5,
                review_status="pending",
                search_status="pending",
                source_sheet="Main",
                sizes=["42"],
            )
        )
        db.commit()
        session_id = sess.id
    finally:
        db.close()

    page = client.get(f"/generate/{session_id}")
    assert page.status_code == 200
    assert "1 approved items" in page.text

    response = client.post(f"/generate/{session_id}", json={"save_images": False})
    assert response.status_code == 200
    assert response.json()["ok"] is True

    excel_file = None
    for _ in range(40):
        poll_db = test_app["database"].SessionLocal()
        try:
            excel_file = (
                poll_db.query(models.GeneratedFile)
                .filter(
                    models.GeneratedFile.session_id == session_id,
                    models.GeneratedFile.filename != "images.zip",
                )
                .first()
            )
            if excel_file is not None:
                poll_db.expunge(excel_file)
                break
        finally:
            poll_db.close()
        time.sleep(0.25)

    assert excel_file is not None
    download = client.get(f"/download/{excel_file.token}")
    assert download.status_code == 200
