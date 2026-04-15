from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.core.generator import OrderSheetGenerator


def _load_headers(path: Path) -> list[str]:
    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(row=2, column=idx).value for idx in range(1, ws.max_column + 1)]
    wb.close()
    return headers


def _make_item(item_code: str, comming_soon_qty=..., size: str = "onesize") -> dict:
    item = {
        "item_code": item_code,
        "style_name": "Test Product",
        "color_name": "Black",
        "color_code": "BLK",
        "gender": "Men",
        "wholesale_price": 10,
        "retail_price": 20,
        "qty_available": 5,
        "size": size,
        "approved_url": "",
        "pictures_url": "",
        "additional_urls": [],
        "brand": "Test Brand",
        "barcode": "1234567890",
        "item_group": "Accessories",
    }
    if comming_soon_qty is not ...:
        item["comming_soon_qty"] = comming_soon_qty
    return item


def test_generator_includes_comming_soon_column_and_exact_values(tmp_path):
    generator = OrderSheetGenerator()
    output = generator.generate(
        items=[
            _make_item("SKU-001", "0"),
            _make_item("SKU-002", "3"),
        ],
        output_dir=str(tmp_path),
        input_filename="reorder_dubai",
        brand="Dubai",
    )

    wb = load_workbook(output)
    ws = wb.active
    headers = _load_headers(Path(output))
    assert "Comming Soon" in headers
    col = headers.index("Comming Soon") + 1
    assert ws.cell(row=3, column=col).value == 0
    assert ws.cell(row=4, column=col).value == 3
    wb.close()


def test_generator_omits_comming_soon_column_when_source_has_no_value(tmp_path):
    generator = OrderSheetGenerator()
    output = generator.generate(
        items=[
            _make_item("SKU-001"),
            _make_item("SKU-002"),
        ],
        output_dir=str(tmp_path),
        input_filename="standard_export",
        brand="FLENDER",
    )

    headers = _load_headers(Path(output))
    assert "Comming Soon" not in headers
