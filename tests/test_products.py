from __future__ import annotations

import openpyxl

from app.core.attribute_engine import build_upload_workbook, parse_sap_products
from app.core.attribute_taxonomy import master_for_item_group


def _make_sap_export(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Code", "Web Description 2", "Item Group", "Material", "Style Code", "Gender"])
    # Two size rows for one style (must dedupe) + a second style
    ws.append(["TNT A 1 BLACK M", "Logo Tee", "T-SHIRTS", "Cotton 100%", "TN001", "Men"])
    ws.append(["TNT A 1 BLACK L", "Logo Tee", "T-SHIRTS", "Cotton 100%", "TN001", "Men"])
    ws.append(["TNT P 2 NAVY M", "Relaxed Jeans", "PANTS", "Cotton 100%", "TN002", "Men"])
    wb.save(path)


def test_master_for_item_group_rolls_up_subgroups():
    assert master_for_item_group("HOODYS") == "SWEATS"
    assert master_for_item_group("HEAD") == "ACCS"
    assert master_for_item_group("LONGSL") == "T-SHIRTS"
    assert master_for_item_group("PANTS") == "PANTS"
    assert master_for_item_group("") == "ACCS"  # safe default


def test_parse_sap_products_dedupes_and_maps(tmp_path):
    p = tmp_path / "export.xlsx"
    _make_sap_export(p)
    styles, meta = parse_sap_products(str(p))
    assert {s["style_code"] for s in styles} == {"TN001", "TN002"}  # deduped
    by_code = {s["style_code"]: s for s in styles}
    assert by_code["TN001"]["master_group"] == "T-SHIRTS"
    assert by_code["TN001"]["name"] == "Logo Tee"
    assert by_code["TN002"]["master_group"] == "PANTS"
    assert "style_code" in meta["columns_found"]


def test_parse_reads_vendor_category_and_long_description(tmp_path):
    """Hamid's Carhartt-style export: 'SAP Item Group' is the real group (not
    'Item Group Code', which holds the full item code), and Vendor Category /
    µ-separated Long Description / Season are captured for the AI."""
    p = tmp_path / "carhartt.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Item Group Code", "Style Code", "Mfr Catalog No", "Web Description",
               "Web Description 2", "SAP Item Group", "Vendor Category",
               "Long Description", "Season"])
    ws.append(["CAR A I000149 4FCXX", "I000149", "I000149.4FCXX", "CARHARTT WIP",
               "C-Logo Sticker (30 Pack)", "ACCS", "POS Divers",
               "8 x 8 cm µ 30 Pack µ screen print", "2027-01"])
    wb.save(p)

    styles, meta = parse_sap_products(str(p))
    assert len(styles) == 1
    s = styles[0]
    assert s["style_code"] == "I000149"
    assert s["item_group"] == "ACCS"          # SAP Item Group wins
    assert s["master_group"] == "ACCS"
    assert s["name"] == "C-Logo Sticker (30 Pack)"
    assert s["vendor_category"] == "POS Divers"
    assert s["long_description"] == "8 x 8 cm; 30 Pack; screen print"  # µ split
    assert s["season"] == "2027-01"
    assert "vendor_category" in meta["columns_found"]


def test_enrich_prompt_carries_catalog_context(monkeypatch):
    """Vendor category, long description and the matched catalog extract must
    all reach the AI prompt."""
    import app.core.attribute_engine as eng
    seen = {}

    def fake_ai(prompt, max_tokens=300):
        seen["prompt"] = prompt
        return '{"product_type":"TSHIRT","confidence":0.9,"FABRIC":null,"FIT":null,"STYLE":[],"WEIGHT":null}'
    monkeypatch.setattr(eng, "_call_ai", fake_ai)

    r = eng.enrich_style({
        "style_code": "TN001", "name": "Logo Tee", "item_group": "T-SHIRTS",
        "master_group": "T-SHIRTS", "material": "Cotton", "gender": "Men",
        "vendor_category": "Graphic Tees", "season": "2027-01",
        "long_description": "240 gsm jersey; boxy fit",
        "ref_text": "Heavyweight tee with dropped shoulders",
    })
    assert r["product_type"] == "TSHIRT"
    assert "vendor_category=Graphic Tees" in seen["prompt"]
    assert "240 gsm jersey" in seen["prompt"]
    assert "Heavyweight tee with dropped shoulders" in seen["prompt"]


def test_build_upload_workbook_drops_review_and_emits_sap_rows(tmp_path):
    results = [
        {"style_code": "TN001", "master_group": "T-SHIRTS", "product_type": "TSHIRT",
         "needs_review": False, "FABRIC": None, "FIT": "Loose", "WEIGHT": "Light",
         "STYLE": ["Street"]},
        {"style_code": "TN002", "master_group": "PANTS", "product_type": None,
         "needs_review": True, "FABRIC": None, "FIT": None, "WEIGHT": None, "STYLE": []},
    ]
    out = tmp_path / "upload.xlsx"
    summary = build_upload_workbook(results, str(out))
    assert summary == {"clean_styles": 1, "rows": 4, "review_styles": 1}

    wb = openpyxl.load_workbook(out)
    ws = wb["Sheet1"]
    rows = [tuple(ws.cell(r, c).value for c in range(1, 5)) for r in range(2, ws.max_row + 1)]
    assert ("TN001", "T-SHIRTS", "TSHIRT", "Y") in rows       # product type Boolean
    assert ("TN001", "T-SHIRTS", "FIT", "Loose") in rows      # valued attribute
    assert ("TN001", "T-SHIRTS", "STYLE", "Street") in rows
    assert all(r[0] != "TN002" for r in rows)                 # review style excluded


def test_products_page_requires_login_and_renders(client, login_as):
    # Unauthenticated -> redirected to login
    anon = client.get("/products", follow_redirects=False)
    assert anon.status_code in (302, 307)

    login_as()
    page = client.get("/products")
    assert page.status_code == 200
    assert "Product Attributes" in page.text


# ── Saved runs / multi-file / editing (Product Attributes tool) ───────────────

def _make_run(db_session, models, uid, results, name="Run", status="done"):
    from app.routers.products_routes import _summary_counts
    s = _summary_counts(results)
    run = models.ProductAttributeRun(
        user_id=uid, name=name, status=status, filename=name,
        total_styles=len(results), clean_count=s["clean_styles"],
        review_count=s["review_styles"], row_count=s["rows"],
    )
    run.results = results
    db_session.add(run)
    db_session.commit()
    db_session.refresh(run)
    return run


def _tee(code="TN001", group="T-SHIRTS", pt="TSHIRT", needs_review=False):
    return {"style_code": code, "name": "Logo Tee", "master_group": group,
            "item_group": group, "product_type": pt, "confidence": 0.9,
            "FABRIC": None, "FIT": None, "WEIGHT": "Light", "STYLE": ["Street"],
            "needs_review": needs_review}


def test_run_persists_and_dedupes_multiple_files(client, login_as, test_app, monkeypatch):
    import app.routers.products_routes as pr
    # Deterministic stub (no AI): first valid product type for the group.
    from app.core.attribute_taxonomy import PRODUCT_TYPES_BY_GROUP

    def fake_enrich(style):
        cands = PRODUCT_TYPES_BY_GROUP.get(style["master_group"], [])
        pt = cands[0][0] if cands else None
        return {**style, "product_type": pt, "confidence": 0.9,
                "FABRIC": None, "FIT": None, "WEIGHT": None, "STYLE": [],
                "needs_review": pt is None}
    monkeypatch.setattr(pr, "enrich_style", fake_enrich)

    user = login_as()
    import io
    def sap_bytes(rows):
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["Web Description 2", "Item Group", "Material", "Style Code", "Gender"])
        for r in rows:
            ws.append(r)
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    f1 = sap_bytes([["Logo Tee", "T-SHIRTS", "Cotton", "TN001", "Men"],
                    ["Cargo", "PANTS", "Cotton", "TN002", "Men"]])
    f2 = sap_bytes([["Logo Tee", "T-SHIRTS", "Cotton", "TN001", "Men"],   # dup style
                    ["Cap", "ACCS", "Cotton", "TN003", "Men"]])
    resp = client.post("/products/run", files=[
        ("file", ("a.xlsx", f1, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ("file", ("b.xlsx", f2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
    ])
    assert resp.status_code == 200, resp.text
    run_id = resp.json()["run_id"]
    assert resp.json()["total"] == 3   # TN001 deduped across the two files

    import time
    for _ in range(50):
        st = client.get(f"/products/run/{run_id}/status").json()
        if st["status"] == "done":
            break
        time.sleep(0.1)
    assert st["status"] == "done"
    assert {p["style_code"] for p in st["preview"]} == {"TN001", "TN002", "TN003"}

    # Persisted and visible in history.
    models = test_app["models"]
    db = test_app["database"].SessionLocal()
    try:
        run = db.get(models.ProductAttributeRun, run_id)
        assert run.status == "done"
        assert run.total_styles == 3
    finally:
        db.close()
    assert "a.xlsx" in client.get("/products").text


def test_run_with_reference_file_feeds_catalog_text_to_enrichment(client, login_as, monkeypatch, test_app):
    """A reference Excel uploaded next to the SAP export: matched styles carry
    ref_text into enrichment, get the has_reference flag in the preview, and
    the bulky catalog text is NOT persisted in the run results."""
    import io
    import time

    import app.routers.products_routes as pr

    seen_styles = {}

    def fake_enrich(style):
        seen_styles[style["style_code"]] = dict(style)
        return {**style, "product_type": "TSHIRT", "confidence": 0.9,
                "FABRIC": None, "FIT": None, "WEIGHT": None, "STYLE": [],
                "needs_review": False}
    monkeypatch.setattr(pr, "enrich_style", fake_enrich)

    def xlsx_bytes(header, rows):
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(header)
        for r in rows:
            ws.append(r)
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    login_as()
    sap = xlsx_bytes(["Web Description 2", "Item Group", "Material", "Style Code", "Gender"],
                     [["Logo Tee", "T-SHIRTS", "Cotton", "TN001", "Men"],
                      ["Cargo Pant", "PANTS", "Cotton", "TN002", "Men"]])
    catalog = xlsx_bytes(["Style", "Details"],
                         [["TN001", "Heavy 240gsm jersey, boxy fit"]])
    resp = client.post("/products/run", files=[
        ("file", ("sap.xlsx", sap, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ("reference", ("catalog.xlsx", catalog, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
    ])
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["ref_matched"] == 1
    assert body["ref_files"] == [{"name": "catalog.xlsx", "matched": 1}]

    for _ in range(50):
        st = client.get(f"/products/run/{body['run_id']}/status").json()
        if st["status"] == "done":
            break
        time.sleep(0.1)
    assert st["status"] == "done"

    # The matched style saw the catalog text; the unmatched one didn't.
    assert "Heavy 240gsm jersey" in seen_styles["TN001"].get("ref_text", "")
    assert "ref_text" not in seen_styles["TN002"]

    by_code = {p["style_code"]: p for p in st["preview"]}
    assert by_code["TN001"]["has_reference"] is True
    assert by_code["TN002"]["has_reference"] is False

    # Persisted results keep the flag but drop the raw catalog text.
    models = test_app["models"]
    db = test_app["database"].SessionLocal()
    try:
        run = db.get(models.ProductAttributeRun, body["run_id"])
        stored = {r["style_code"]: r for r in run.results}
        assert stored["TN001"].get("has_reference") is True
        assert "ref_text" not in stored["TN001"]
    finally:
        db.close()


def test_run_rejects_bad_reference_extension(client, login_as):
    import io
    login_as()
    wb = openpyxl.Workbook(); wb.active.append(["Style Code"]); wb.active.append(["TN001"])
    buf = io.BytesIO(); wb.save(buf)
    resp = client.post("/products/run", files=[
        ("file", ("sap.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ("reference", ("notes.txt", b"hello", "text/plain")),
    ])
    assert resp.status_code == 400
    assert "notes.txt" in resp.json()["error"]


def test_edit_style_changes_type_and_clears_review(client, login_as, db_session, test_app):
    models = test_app["models"]
    user = login_as()
    run = _make_run(db_session, models, user["id"], [
        _tee("TN001", "ACCS", pt=None, needs_review=True),  # flagged, no type
    ])
    assert run.review_count == 1

    resp = client.post(f"/products/run/{run.id}/style",
                       json={"style_code": "TN001", "product_type": "BAG", "WEIGHT": "Heavy"})
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["row"]["product_type"] == "BAG"
    assert body["row"]["needs_review"] is False
    assert body["summary"]["review_styles"] == 0

    db_session.expire_all()
    fresh = db_session.get(models.ProductAttributeRun, run.id)
    r0 = fresh.results[0]
    assert r0["product_type"] == "BAG"
    assert r0["WEIGHT"] == "Heavy"
    assert r0["edited"] is True
    assert fresh.review_count == 0


def test_edit_rejects_invalid_product_type_for_group(client, login_as, db_session, test_app):
    models = test_app["models"]
    user = login_as()
    run = _make_run(db_session, models, user["id"], [_tee("TN001", "ACCS", pt=None, needs_review=True)])
    # SNEAKER is a SHOES type, not valid for an ACCS style.
    resp = client.post(f"/products/run/{run.id}/style",
                       json={"style_code": "TN001", "product_type": "SNEAKER"})
    assert resp.status_code == 400


def test_reopen_download_and_delete(client, login_as, db_session, test_app):
    models = test_app["models"]
    user = login_as()
    run = _make_run(db_session, models, user["id"], [_tee("TN001")])

    # Reopen
    d = client.get(f"/products/run/{run.id}").json()
    assert d["preview"][0]["style_code"] == "TN001"

    # Download reflects the saved results
    dl = client.get(f"/products/download/{run.id}")
    assert dl.status_code == 200
    wb = openpyxl.load_workbook(io_bytes(dl.content))
    codes = {wb["Sheet1"].cell(r, 1).value for r in range(2, wb["Sheet1"].max_row + 1)}
    assert "TN001" in codes

    # Delete
    assert client.post(f"/products/run/{run.id}/delete").status_code == 200
    db2 = test_app["database"].SessionLocal()
    try:
        assert db2.get(models.ProductAttributeRun, run.id) is None
    finally:
        db2.close()


def test_run_ownership_enforced(client, login_as, db_session, test_app, make_user):
    models = test_app["models"]
    owner = make_user(username="po", email="po@flendergroup.com")
    run = _make_run(db_session, models, owner["id"], [_tee("TN001")])
    login_as(username="po_intruder", email="po_intruder@flendergroup.com")
    assert client.get(f"/products/run/{run.id}").status_code == 404
    assert client.post(f"/products/run/{run.id}/style",
                      json={"style_code": "TN001", "product_type": "TSHIRT"}).status_code == 404


def io_bytes(content):
    import io
    return io.BytesIO(content)
