"""Full end-to-end probe: searcher + AI VISION ranking on the real sheet.

For each sampled product it runs the live searcher, then the vision ranker
(which fetches the actual images and hard-rejects wrong colour / wrong product
/ detail crops / lifestyle shots), and reports the chosen image — the true
"would Jane keep this?" signal, not just source quality."""
import os, re, sys, random
from pathlib import Path
from dotenv import load_dotenv
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")
sys.path.insert(0, str(ROOT))
from app.core.searcher import ImageSearcher                      # noqa: E402
from app.services.ai_service import ai_rank_urls                 # noqa: E402

SHEET = sys.argv[1] if len(sys.argv) > 1 else "/Users/kareemelsenosy/Downloads/Ordersheet Test Drole.xlsx"
N = int(sys.argv[2]) if len(sys.argv) > 2 else 10

ws = openpyxl.load_workbook(SHEET)["List"]
seen, items = set(), []
for r in range(2, ws.max_row + 1):
    b, code, desc, col = (ws.cell(r, c).value for c in (1, 2, 3, 6))
    if not code:
        continue
    key = (str(desc).strip().lower(), str(col).strip().lower())
    if key in seen:
        continue
    seen.add(key)
    items.append(dict(brand=str(b), item_code=str(code),
                      style_name=str(desc).strip(), color_name=str(col).strip()))
random.seed(7)
sample = random.sample(items, min(N, len(items)))

s = ImageSearcher({"google_api_key": os.getenv("GOOGLE_SEARCH_KEY", ""),
                   "google_cse_id": os.getenv("GOOGLE_CSE_ID", "")})
usable = 0
print(f"{'STYLE / COLOR':40} {'#cand':5} {'#kept':5} {'CHOSEN IMAGE DOMAIN':24} verdict")
print("-" * 96)
for it in sample:
    cands, scores = s.search(it)
    if not cands:
        print(f"{(it['style_name'][:26]+' / '+it['color_name'])[:40]:40} {'0':5} {'-':5} {'(search empty)':24} BLANK")
        continue
    ranked, discarded = ai_rank_urls(cands, it, it["brand"], scores=scores, prefer_vision=True)
    kept = [u for u in ranked if u not in discarded]
    chosen = kept[0] if kept else ""
    dom = re.sub(r"^https?://(www\.)?", "", chosen).split("/")[0] if chosen else "(all rejected)"
    if chosen:
        usable += 1
        verdict = "USABLE"
    else:
        verdict = "BLANK (vision rejected all)"
    label = (it['style_name'][:26] + ' / ' + it['color_name'])[:40]
    print(f"{label:40} {len(cands):<5} {len(kept):<5} {dom[:24]:24} {verdict}")
print("-" * 96)
print(f"Products with a vision-approved usable image: {usable}/{len(sample)} = {round(100*usable/len(sample))}%")
print("(BLANK = vision rejected everything -> empty cell, which is correct per 'empty beats wrong')")
