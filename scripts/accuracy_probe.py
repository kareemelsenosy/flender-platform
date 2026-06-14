"""One-off accuracy probe for the Drôle de Monsieur order sheet.

Measures how often the searcher's top image candidate lands on the brand site
or a trusted stockist with the right colour — an automated proxy for "usable
image" (the AI-vision re-ranker is a separate downstream layer)."""
import os, re, random, sys
from pathlib import Path
from dotenv import load_dotenv
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")
sys.path.insert(0, str(ROOT))
from app.core.searcher import ImageSearcher, _tokenize  # noqa: E402

SHEET = sys.argv[1] if len(sys.argv) > 1 else "/Users/kareemelsenosy/Downloads/Ordersheet Test Drole.xlsx"
N = int(sys.argv[2]) if len(sys.argv) > 2 else 12

ws = openpyxl.load_workbook(SHEET)["List"]
seen, items = set(), []
for r in range(2, ws.max_row + 1):
    b, code, desc, col = (ws.cell(r, c).value for c in (1, 2, 3, 6))
    if not code:
        continue
    key = (str(desc).strip().lower(), str(col).strip().lower())
    if key in seen:
        continue
    seen.add(key)
    items.append(dict(brand=str(b), item_code=str(code),
                      style_name=str(desc).strip(), color_name=str(col).strip()))
random.seed(7)
sample = random.sample(items, min(N, len(items)))

s = ImageSearcher({"google_api_key": os.getenv("GOOGLE_SEARCH_KEY", ""),
                   "google_cse_id": os.getenv("GOOGLE_CSE_ID", "")})
# Known legit fashion-retail / stockist hosts (incl. image-CDN hostnames that
# sit under a different domain than the storefront, e.g. ssensemedia/lystit).
STOCKIST = ("droledemonsieur", "ssense", "ssensemedia", "wrongweather", "modesens",
            "farfetch", "lyst", "lystit", "endclothing", "slamjam", "antonioli",
            "bstn", "mintcompany", "michaelchell", "clothbase", "stormfashion",
            "outbacksylt", "cloudinary", "shopify", "selfridges", "mytheresa",
            "ssensecdn", "nakedcph", "voo-store", "hbx", "luisaviaroma")
JUNK = ("ebay", "amazon", "aliexpress", "pinterest", "pinimg", "temu", "dhgate",
        "wish.", "vinted", "poshmark", "depop", "grailed", "alibaba")
stockist_hits = junk_hits = 0
print(f"{'STYLE / COLOR':42} {'PAGE DOMAIN':24} {'score':5} class")
print("-" * 92)
for it in sample:
    try:
        cands, scores = s.search(it)
    except Exception as e:
        print(f"{(it['style_name'][:28]+' / '+it['color_name'])[:42]:42} ERROR {e}")
        continue
    top = cands[0] if cands else ""
    idom = re.sub(r"^https?://(www\.)?", "", top).split("/")[0] if top else "(none)"
    blob = top.lower()
    if any(j in blob for j in JUNK):
        cls = "JUNK"; junk_hits += 1
    elif any(t in blob for t in STOCKIST):
        cls = "stockist/retail"; stockist_hits += 1
    elif top:
        cls = "other"
    else:
        cls = "MISS"
    label = (it['style_name'][:28] + ' / ' + it['color_name'])[:42]
    sc = scores.get(top, 0.0) if top else 0.0
    print(f"{label:42} {idom[:24]:24} {sc:<5.2f} {cls}")
print("-" * 92)
print(f"Top result on a legit stockist/retailer: {stockist_hits}/{len(sample)} = {round(100*stockist_hits/len(sample))}%")
print(f"Top result on marketplace/junk        : {junk_hits}/{len(sample)}")
